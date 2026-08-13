#!/usr/bin/env python3
"""25년 하반기 개인별 성과급.xlsx → bonus_statements(2025, H2) 임포트 (일회성, 블루프린트 §8).

1) '성과급 총괄표' 시트 우측 '전직원 성과급 현황'(H=성명, I=직급, J=지급액, 23행~)을
   직원 매칭(이름, 동명이인은 직급으로 재판별)해 적재한다. 지급액 > 0 만 적재.
   26년 상반기 명세 화면의 '전기 성과급' 열 근거 데이터.
2) 개인 시트(시트명=성명, A=연번/B=용역명/G=매출기여금액/H=참여도(0~1)/I=평점(구 0~5)/J=성과산정액1)를
   bonus_statement_lines 로 적재한다(지급 명세서의 용역별 내역표 근거).
   계약명 완전일치 매칭 성공분만 적재, 실패는 리포트(재실행 시 lines 전량 재생성 = 멱등).

사용법:
  DATABASE_URL=postgresql://... python scripts/import_bonus_2025h2.py --dry-run
  DATABASE_URL=postgresql://... python scripts/import_bonus_2025h2.py
"""
from __future__ import annotations

import argparse
import json
import os
import re
import secrets
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
import psycopg

DEFAULT_XLSX = Path(
    r"V:\업무자료\경리업무\성과급 관련\성과급 지급현황\2025\하반기\25년 하반기 개인별 성과급.xlsx"
)
SHEET = "성과급 총괄표"
C_NAME, C_POSITION, C_AMOUNT = 8, 9, 10  # H, I, J
DATA_START = 23


def norm(s) -> str:
    return re.sub(r"\s+", " ", (str(s) if s is not None else "").strip())


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    dsn = os.environ.get("DATABASE_URL")
    if not dsn:
        raise SystemExit("DATABASE_URL 환경변수가 필요합니다.")

    wb = openpyxl.load_workbook(args.xlsx, data_only=True, read_only=True)
    ws = wb[SHEET]
    rows = []
    for r in ws.iter_rows(min_row=DATA_START, max_col=C_AMOUNT, values_only=True):
        name = norm(r[C_NAME - 1])
        if not name:
            continue
        position = norm(r[C_POSITION - 1])
        try:
            amount = float(r[C_AMOUNT - 1] or 0)
        except (TypeError, ValueError):
            amount = 0.0
        if amount > 0:
            rows.append({"name": name, "position": position, "amount": amount})
    print(f"엑셀 지급액>0 인원: {len(rows)}명, 합계 {sum(x['amount'] for x in rows):,.0f}원")

    # 개인 시트 → 용역별 내역 (B=용역명, G=기여금액, H=참여도(0~1), I=평점(구 0~5), J=산정액1)
    def num(v) -> float:
        try:
            return float(v or 0)
        except (TypeError, ValueError):
            return 0.0

    lines_by_name: dict[str, list[dict]] = {}
    for row in rows:
        if row["name"] not in wb.sheetnames:
            continue
        pws = wb[row["name"]]
        items = []
        for r in pws.iter_rows(min_row=7, max_col=10, values_only=True):
            title = norm(r[1])
            amount = num(r[9])
            if not title or title == "0" or amount <= 0:
                continue
            items.append({
                "title": title,
                "applied": num(r[6]),
                "pct": num(r[7]) * 100,
                "grade": str(int(num(r[8]))) if num(r[8]) > 0 else None,
                "amount": amount,
            })
        if items:
            lines_by_name[row["name"]] = items
    print(f"개인 시트 용역별 내역 보유: {len(lines_by_name)}명 / {sum(len(v) for v in lines_by_name.values())}건")
    wb.close()

    now = datetime.now(timezone.utc).isoformat()
    meta = json.dumps({"source": "excel-25h2", "note": "25년 하반기 개인별 성과급.xlsx 전직원 현황 임포트"})
    stats = {"insert": 0, "no_match": 0, "ambiguous": 0, "line_insert": 0, "line_no_match": 0}
    report = []

    with psycopg.connect(dsn) as conn:
        with conn.cursor() as cur:
            cur.execute(
                """SELECT e.employee_id, e.name, COALESCE(po.position_name, '')
                     FROM employee_profiles e
                     LEFT JOIN positions po ON po.position_id = e.position_id"""
            )
            by_name: dict[str, list[tuple[str, str]]] = {}
            for eid, name, pos in cur.fetchall():
                by_name.setdefault(norm(name), []).append((eid, norm(pos)))

            cur.execute("SELECT contract_id, contract_title FROM contracts WHERE deleted_at IS NULL")
            contract_by_title: dict[str, list[str]] = {}
            for cid, title in cur.fetchall():
                contract_by_title.setdefault(norm(title), []).append(cid)

            for row in rows:
                candidates = by_name.get(row["name"], [])
                if not candidates:
                    stats["no_match"] += 1
                    report.append(f"[매칭실패] {row['name']}({row['position']}) {row['amount']:,.0f}원")
                    continue
                if len(candidates) > 1:
                    narrowed = [c for c in candidates if c[1] == row["position"]]
                    if len(narrowed) != 1:
                        stats["ambiguous"] += 1
                        report.append(f"[동명이인 판별불가] {row['name']}({row['position']}) — 후보 {len(candidates)}명, 스킵")
                        continue
                    candidates = narrowed
                eid = candidates[0][0]
                stats["insert"] += 1
                statement_id = None
                if not args.dry_run:
                    cur.execute(
                        """INSERT INTO bonus_statements
                             (statement_id, period_year, period_half, employee_id, bucket,
                              total_amount, calc_method, meta_json, calculated_at)
                           VALUES (%s, 2025, 'H2', %s, 'participant', %s, 'base', %s::jsonb, %s)
                           ON CONFLICT (period_year, period_half, employee_id) DO UPDATE SET
                             total_amount = EXCLUDED.total_amount,
                             meta_json = EXCLUDED.meta_json,
                             calculated_at = EXCLUDED.calculated_at
                           RETURNING statement_id""",
                        (f"bst_{secrets.token_hex(8)}", eid, row["amount"], meta, now),
                    )
                    statement_id = cur.fetchone()[0]

                # 개인 시트 용역별 내역 → lines (계약명 완전일치만, 전량 재생성)
                items = lines_by_name.get(row["name"], [])
                if items and not args.dry_run:
                    cur.execute("DELETE FROM bonus_statement_lines WHERE statement_id = %s", (statement_id,))
                for item in items:
                    cids = contract_by_title.get(item["title"], [])
                    cid = cids[0] if len(cids) == 1 else None
                    if cid is None:
                        # 계약 미매칭(엑셀 집계 행·미등록 계약)도 제목 스냅샷으로 적재해 합계를 보존한다
                        stats["line_no_match"] += 1
                        report.append(
                            f"[내역 계약매칭 실패→제목 적재] {row['name']}: {item['title']} ({'중복' if cids else '없음'})"
                        )
                    stats["line_insert"] += 1
                    if not args.dry_run:
                        cur.execute(
                            """INSERT INTO bonus_statement_lines
                                 (line_id, statement_id, contract_id, contract_title, applied_amount,
                                  participation_pct, grade, amount)
                               VALUES (%s, %s, %s, %s, %s, %s, %s, %s)""",
                            (f"bsl_{secrets.token_hex(8)}", statement_id, cid, item["title"],
                             item["applied"], item["pct"], item["grade"], item["amount"]),
                        )
        if args.dry_run:
            conn.rollback()
        else:
            conn.commit()

    print("--- 리포트 ---")
    for line in report:
        print(line)
    mode = "DRY-RUN" if args.dry_run else "적용"
    print(
        f"[{mode}] 적재 {stats['insert']} / 매칭실패 {stats['no_match']} / 동명이인 스킵 {stats['ambiguous']} / "
        f"내역 적재 {stats['line_insert']} / 내역 매칭실패 {stats['line_no_match']}"
    )


if __name__ == "__main__":
    main()
