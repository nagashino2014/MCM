#!/usr/bin/env python3
"""사업참여수행인력 현황 엑셀 → MCM DB 마이그레이션 (일회성).

두 가지를 수행한다.

1) 인원 속성 백필 (입사일 시트, A열 성명으로 employee_profiles 매칭)
   - agent_registered_at(대행인력등록일), eng_grade(기술등급), specialty_field(전문분야) : 엑셀이 권위 → 직접 갱신
   - birth_date, job_duties : 기존 값이 없을 때만 채움(COALESCE)
   - env_grade(환경부 고급/일반) : 건드리지 않음 — 수동 입력 항목
   - 조직(employee_profiles)에 이름이 없는 인원(퇴사자)은 무시

2) 수행인력 적재 (수행인력 현황 시트, B열 용역명으로 contracts.contract_title 매칭)
   - 본부(O열) → departments.dept_name 매칭 → contracts.owning_dept_id (현재 NULL인 경우만 설정)
   - 참여자 = 본부장(N열, role='관리자') + 참여자1~6(H~M열, role='실무'); 동일인은 관리자 우선 1행
   - 조직에 이름이 없는 인원은 스킵(퇴사자)
   - participated_from = max(용역시작일, 직원 대행인력등록일)  ★대행인력등록일 하한(환경부 통합허가 대행업 등록 준수)
   - participated_to  = min(용역마감일['진행중'→NULL], 직원 퇴사일)
   - task_label = 입사일!J(담당업무), field_label = 직원 전문분야
   - service_participants 에 source='manual' 로 ON CONFLICT upsert

사용법:
  DATABASE_URL=postgresql://... python scripts/migrate_staffing_from_excel.py --dry-run
  DATABASE_URL=postgresql://... python scripts/migrate_staffing_from_excel.py
"""
from __future__ import annotations

import argparse
import os
import re
import secrets
from datetime import date, datetime
from pathlib import Path
from typing import Optional

import openpyxl
import psycopg

DEFAULT_XLSX = Path(
    r"C:\CodingProject\사무용유틸\수행인력 자동입력\사업참여수행인력 현황(수행인력 업데이트).xlsx"
)
SERVICE_SHEET = "수행인력 현황"
PERSON_SHEET = "입사일"

# 수행인력 현황 시트 열 (1-indexed). 데이터는 4행부터.
C_SVC_NAME = 2   # B 용역명
C_SVC_START = 6  # F 용역 시작일
C_SVC_END = 7    # G 용역 마감일 ('진행중' 또는 날짜)
C_PART_FIRST = 8  # H 참여자1
C_PART_LAST = 13  # M 참여자6
C_MANAGER = 14   # N 본부장(관리자)
C_DEPT = 15      # O 본부
SVC_DATA_START = 4

# 입사일 시트 열 (1-indexed). 데이터는 3행부터.
P_NAME = 1       # A 성명
P_BIRTH = 5      # E 생년월일
P_AGENT_REG = 6  # F 대행인력등록일
P_RESIGN = 7     # G 퇴사일
P_ENG_GRADE = 8  # H 기술등급(엔지니어링)
P_SPECIALTY = 9  # I 전문분야
P_JOB = 10       # J 담당업무
PERSON_DATA_START = 3


def parse_date(v) -> Optional[date]:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    if not s:
        return None
    s = s.replace("년", ".").replace("월", ".").replace("일", "")
    s = s.replace("/", ".").replace("-", ".")
    m = re.search(r"(\d{4})\D?(\d{1,2})\D?(\d{1,2})", s)
    if m:
        try:
            return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            return None
    digits = re.sub(r"\D", "", s)
    if len(digits) == 8:
        try:
            return date(int(digits[:4]), int(digits[4:6]), int(digits[6:8]))
        except ValueError:
            return None
    return None


def iso(d: Optional[date]) -> Optional[str]:
    return d.isoformat() if d else None


def norm(s: Optional[str]) -> str:
    """매칭용 정규화: 양끝 공백 제거 + 내부 연속 공백 1칸."""
    return re.sub(r"\s+", " ", (s or "").strip())


def cell_str(ws, row: int, col: int) -> str:
    v = ws.cell(row=row, column=col).value
    return str(v).strip() if v is not None else ""


def main() -> int:
    ap = argparse.ArgumentParser(description="수행인력 엑셀 → MCM DB 마이그레이션")
    ap.add_argument("--xlsx", default=str(DEFAULT_XLSX))
    ap.add_argument("--database-url", default=os.environ.get("DATABASE_URL"))
    ap.add_argument("--dry-run", action="store_true", help="DB 변경 없이 계획만 출력")
    args = ap.parse_args()

    if not args.database_url:
        raise SystemExit("DATABASE_URL 또는 --database-url 이 필요합니다.")
    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        raise SystemExit(f"엑셀을 찾을 수 없습니다: {xlsx_path}")

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws_svc = wb[SERVICE_SHEET]
    ws_per = wb[PERSON_SHEET]

    # 입사일 시트: 이름 → 속성
    person_attrs: dict[str, dict] = {}
    for r in range(PERSON_DATA_START, ws_per.max_row + 1):
        name = norm(cell_str(ws_per, r, P_NAME))
        if not name:
            continue
        person_attrs[name] = {
            "birth": parse_date(ws_per.cell(row=r, column=P_BIRTH).value),
            "agent_reg": parse_date(ws_per.cell(row=r, column=P_AGENT_REG).value),
            "resign": parse_date(ws_per.cell(row=r, column=P_RESIGN).value),
            "eng_grade": cell_str(ws_per, r, P_ENG_GRADE) or None,
            "specialty": cell_str(ws_per, r, P_SPECIALTY) or None,
            "job": cell_str(ws_per, r, P_JOB) or None,
        }

    now = datetime.now().isoformat()
    conn = psycopg.connect(args.database_url, connect_timeout=10)
    try:
        cur = conn.cursor()

        # 조직 명단 (이름 → employee_id). 동명이인 경고.
        cur.execute("SELECT employee_id, name FROM employee_profiles")
        roster: dict[str, list[str]] = {}
        for emp_id, name in cur.fetchall():
            roster.setdefault(norm(name), []).append(emp_id)
        dups = {n: ids for n, ids in roster.items() if len(ids) > 1}
        for n, ids in dups.items():
            print(f"[경고] 동명이인 '{n}' {len(ids)}명 → 첫 번째 사용")

        # 부서명 → dept_id
        cur.execute("SELECT dept_id, dept_name FROM departments")
        dept_by_name = {norm(name): dept_id for dept_id, name in cur.fetchall()}

        # 계약명(정규화) → [contract_id]
        cur.execute("SELECT contract_id, contract_title, owning_dept_id FROM contracts")
        contract_by_title: dict[str, list[str]] = {}
        owning_now: dict[str, Optional[str]] = {}
        for cid, title, owning in cur.fetchall():
            contract_by_title.setdefault(norm(title), []).append(cid)
            owning_now[cid] = owning

        def resolve_emp(name: str) -> Optional[str]:
            ids = roster.get(norm(name))
            return ids[0] if ids else None

        # ---- 1) 인원 속성 백필 ----
        backfilled = 0
        for name, a in person_attrs.items():
            emp_id = resolve_emp(name)
            if not emp_id:
                continue
            if not args.dry_run:
                cur.execute(
                    """
                    UPDATE employee_profiles SET
                        agent_registered_at = %s,
                        eng_grade           = %s,
                        specialty_field     = %s,
                        birth_date  = COALESCE(NULLIF(birth_date, ''), %s),
                        job_duties  = COALESCE(NULLIF(job_duties, ''), %s),
                        updated_at  = %s
                    WHERE employee_id = %s
                    """,
                    [iso(a["agent_reg"]), a["eng_grade"], a["specialty"],
                     iso(a["birth"]), a["job"], now, emp_id],
                )
            backfilled += 1

        # ---- 2) 수행인력 적재 ----
        matched_svc = 0
        unmatched_titles: list[str] = []
        ambiguous_titles: list[str] = []
        skipped_persons: set[str] = set()
        participant_upserts = 0
        owning_set = 0

        for r in range(SVC_DATA_START, ws_svc.max_row + 1):
            svc_name = cell_str(ws_svc, r, C_SVC_NAME)
            if not svc_name:
                continue
            cids = contract_by_title.get(norm(svc_name))
            if not cids:
                unmatched_titles.append(svc_name)
                continue
            if len(cids) > 1:
                ambiguous_titles.append(svc_name)
            contract_id = cids[0]
            matched_svc += 1

            # 본부 → owning_dept_id (현재 NULL일 때만)
            dept_name = cell_str(ws_svc, r, C_DEPT)
            dept_id = dept_by_name.get(norm(dept_name)) if dept_name else None
            if dept_id and not owning_now.get(contract_id):
                if not args.dry_run:
                    cur.execute(
                        "UPDATE contracts SET owning_dept_id = %s, updated_at = %s "
                        "WHERE contract_id = %s AND owning_dept_id IS NULL",
                        [dept_id, now, contract_id],
                    )
                owning_now[contract_id] = dept_id
                owning_set += 1

            svc_start = parse_date(ws_svc.cell(row=r, column=C_SVC_START).value)
            end_raw = cell_str(ws_svc, r, C_SVC_END)
            svc_end = None if end_raw == "진행중" else parse_date(end_raw)

            # 참여자 수집: 동일인은 관리자(N) 우선 1행. emp_id → (role, 엑셀상 성명)
            role_by_emp: dict[str, tuple[str, str]] = {}
            mgr = cell_str(ws_svc, r, C_MANAGER)
            if mgr:
                emp_id = resolve_emp(mgr)
                if emp_id:
                    role_by_emp[emp_id] = ("관리자", mgr)
                else:
                    skipped_persons.add(mgr)
            for col in range(C_PART_FIRST, C_PART_LAST + 1):
                nm = cell_str(ws_svc, r, col)
                if not nm:
                    continue
                emp_id = resolve_emp(nm)
                if not emp_id:
                    skipped_persons.add(nm)
                    continue
                role_by_emp.setdefault(emp_id, ("실무", nm))

            for emp_id, (role, src_name) in role_by_emp.items():
                a = person_attrs.get(norm(src_name), {})
                agent_reg = a.get("agent_reg")
                resign = a.get("resign")
                start = svc_start
                if agent_reg and (start is None or agent_reg > start):
                    start = agent_reg  # ★대행인력등록일 하한
                end = svc_end
                if resign and (end is None or resign < end):
                    end = resign
                if not args.dry_run:
                    cur.execute(
                        """
                        INSERT INTO service_participants
                            (participation_id, contract_id, employee_id, role_label,
                             task_label, field_label, participated_from, participated_to,
                             source, created_at, updated_at)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 'manual', %s, %s)
                        ON CONFLICT (contract_id, employee_id, role_label) DO UPDATE SET
                            task_label        = EXCLUDED.task_label,
                            field_label       = EXCLUDED.field_label,
                            participated_from = EXCLUDED.participated_from,
                            participated_to   = EXCLUDED.participated_to,
                            source            = 'manual',
                            updated_at        = EXCLUDED.updated_at
                        """,
                        ["sp_" + secrets.token_hex(8), contract_id, emp_id, role,
                         a.get("job"), a.get("specialty"), iso(start), iso(end), now, now],
                    )
                participant_upserts += 1

        if args.dry_run:
            conn.rollback()
        else:
            conn.commit()
    finally:
        conn.close()

    # ---- 리포트 ----
    print("\n===== 마이그레이션 " + ("(DRY-RUN, 미적용)" if args.dry_run else "(적용 완료)") + " =====")
    print(f"인원 백필           : {backfilled}명")
    print(f"매칭된 용역         : {matched_svc}건")
    print(f"수행인력 upsert     : {participant_upserts}행")
    print(f"owning_dept 설정    : {owning_set}건")
    print(f"매칭 실패 용역      : {len(unmatched_titles)}건")
    print(f"중복 용역명(첫건사용): {len(ambiguous_titles)}건")
    print(f"조직 외 인원(스킵)  : {len(skipped_persons)}명")
    if unmatched_titles:
        print("\n[매칭 실패 용역 예시 최대 20]")
        for t in unmatched_titles[:20]:
            print("   -", t)
    if skipped_persons:
        print("\n[조직 명단에 없는 인원(퇴사자로 간주, 무시)]")
        print("   " + ", ".join(sorted(skipped_persons)))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
