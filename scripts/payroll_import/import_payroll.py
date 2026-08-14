#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""과거 급여대장 엑셀(2016~2026) → payroll_* 테이블 적재 (PL-P0, 블루프린트 §3-1).

원본: V:\\업무자료\\경리업무\\급상여대장\\<연도>년\\*.xlsx (3단 구조 급상여대장)
  - 헤더 3행(사원번호/성명 · 입사일/직급 · 퇴사일/부서) ↔ 직원당 3행 블록이 1:1 대응
  - 2016년 3~7월은 암호화(암호 1234, §8-1) → msoffcrypto 복호화 후 동일 파싱
  - 지급결의서·경비명세·연봉요약은 제외, 상여대장(명절 등)·인턴대장은 포함(§8-3)

단계: inventory → decrypt → parse → normalize(별칭) → validate(3중 합계) → [match → load]
  - 합계 검증: Σ지급라인=지급합계, Σ공제라인=공제합계, 지급−공제=차인지급액 (불일치는 리포트+entry.note)
  - '비과세계' 등 정보성 합계 열은 라인 생성에서 제외(SKIP_INFO)
  - 미지의 항목명 발견 시 payroll_item_defs 에 자동 추가(kind 는 열 구역으로 추정) + 리포트

사용법:
  python scripts/payroll_import/import_payroll.py --dry-run            # 파싱+검증+리포트만 (DB 불필요)
  DATABASE_URL=postgresql://... python scripts/payroll_import/import_payroll.py   # 적재(검수 게이트 후)
리포트: scripts/payroll_import/PAYROLL_IMPORT_REPORT.md
"""
from __future__ import annotations

import argparse
import hashlib
import io
import json
import os
import re
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Optional

import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

BASE = Path(r"V:\업무자료\경리업무\급상여대장")
REPORT_PATH = Path(__file__).parent / "PAYROLL_IMPORT_REPORT.md"
XLSX_PASSWORD = "1234"  # 2016년 3~7월 암호화 파일 (§8-1)

# 인적사항·합계·정보성 열 (공백 제거 후 비교)
PERSON_LABELS = {"사원번호", "성명", "입사일", "직급", "퇴사일", "부서"}
TOTAL_PAY = "지급합계"
TOTAL_DED = "공제합계"
TOTAL_NET = "차인지급액"
SKIP_INFO = {"비과세계", "비고", "영수인", "확인"}  # 라인 생성 제외(정보성)

# 제외 파일 (§8-3: 지급결의서·경비명세·연봉요약 + 대장이 아닌 메모)
# 설상여금명세 = 2017년 구정 상여 안내 메모(대장 아님, 금액은 17년 2월 대장에 반영)
EXCLUDE_PATTERNS = re.compile(r"(지급결의서|경비명세|직원별 연봉|급여정리|바로 가기|설상여금명세)")

# ── 열 밀림 보정 (사용자 확정 2026-08-13, 블루프린트 §1-1) ─────────────────────
# 최근 대장 6개 파일은 1단 헤더 라벨과 실제 기입 항목이 한 칸씩 어긋나 있다
# (예: '출장숙박수당' 라벨 열에 전 직원 식대 200,000 기입). 2단 이하는 정위치.
# 매핑 근거: 식대(전 직원 동일 금액)·자격증수당(3/6/9만 차등) 신호 + 사용자 실물 확인.
# 21년 10~11월·22년 2~4월의 식대 부재는 미지급 기간(구내식당 식권)으로 보정 대상 아님.
_OVR_2504 = {"상여": "overtime", "연차수당": "trip-lodging", "전월미지급금": "meal",
             "초과근무수당": "cert", "연구수당": "housing", "성과급": "longevity"}
_OVR_2508 = {"육아수당": "overtime", "초과근무수당": "trip-lodging", "출장숙박수당": "meal",
             "식대": "cert", "성과급": "longevity"}
_OVR_2603 = {"상여": "prev-unpaid", "전월미지급금": "overtime", "초과근무수당": "trip-lodging",
             "출장숙박수당": "meal", "식대": "cert", "자격증수당": "housing"}
_OVR_2606 = {"성과급": "trip-lodging", "출장숙박수당": "meal", "식대": "cert",
             "자격증수당": "housing", "주거비지원": "longevity"}
COLUMN_OVERRIDES: dict[str, dict[str, str]] = {
    r"2025년\25년04월 급상여대장.xlsx": _OVR_2504,
    r"2025년\25년08월 급상여대장.xlsx": _OVR_2508,
    r"2026년\26년03월 급상여대장.xlsx": _OVR_2603,
    r"2026년\26년04월 급상여대장.xlsx": _OVR_2603,
    r"2026년\26년06월 급상여대장.xlsx": _OVR_2606,
    r"2026년\26년07월 급상여대장.xlsx": _OVR_2606,
}
# 사람 단위 예외: 26.03/04 의 '자격증수당' 라벨 열은 주거비지원이지만 이태하만 장기근속휴가수당
# (사용자 확인 — 26.07 화면의 이태하 500,000 건과 동일 성격).
NAME_OVERRIDES: dict[tuple[str, str, str], str] = {
    (r"2026년\26년03월 급상여대장.xlsx", "자격증수당", "이태하"): "longevity",
    (r"2026년\26년04월 급상여대장.xlsx", "자격증수당", "이태하"): "longevity",
}


def squash(s) -> str:
    """공백·개행 제거 정규화 (헤더 라벨 비교용)."""
    return re.sub(r"\s+", "", str(s or ""))


def norm_item(s) -> str:
    """항목명 정규화: 양끝 공백 제거 + 내부 연속 공백 1칸."""
    return re.sub(r"\s+", " ", str(s or "").strip())


def norm_date(v) -> Optional[str]:
    """대장의 날짜 표기(datetime, '16. 3. 28', '2016-08-01')를 YYYY-MM-DD 로."""
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    if not s:
        return None
    m = re.match(r"^(\d{2,4})[.\-/년\s]+(\d{1,2})[.\-/월\s]+(\d{1,2})", s)
    if m:
        y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if y < 100:
            y += 2000
        try:
            return date(y, mo, d).strftime("%Y-%m-%d")
        except ValueError:
            return s
    return s


def num(v) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(",", "").strip()
    if not s or s in ("-", "·"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


# ---------------------------------------------------------------- inventory

FNAME_RE = [
    # 26년03월 급상여대장 / 17년08월 상여대장 / 19년07월 상여(하계휴가)대장
    re.compile(r"^(\d{2})년\s*(\d{2})월\s*(급여|급상여|상여)(?:\((?P<tag>[^)]*)\))?대장"),
    # 환안연 3월 급여대장 / 환안연 4월급여대장 (연도는 폴더에서)
    re.compile(r"^환안연\s*(?P<m_only>\d{1,2})월\s*급여대장"),
    # 인턴급여대장(19.01.31)
    re.compile(r"^인턴급여대장\((?P<intern_date>[\d.]+)\)"),
]


def inventory() -> tuple[list[dict], list[str]]:
    """대상 파일 수집 + (연/월/종류) 파일명 파싱. return (targets, excluded)."""
    targets, excluded = [], []
    files = sorted(BASE.glob("*년/*.xlsx")) + sorted(BASE.glob("인턴급여대장*.xlsx"))
    for p in files:
        rel = str(p.relative_to(BASE))
        if EXCLUDE_PATTERNS.search(p.name):
            excluded.append(f"{rel} — 제외 패턴")
            continue
        name = p.stem
        year = month = None
        kind = "salary"
        folder_year = re.match(r"^(\d{4})년", p.parent.name)
        m = FNAME_RE[0].match(name)
        if m:
            year, month = 2000 + int(m.group(1)), int(m.group(2))
            if m.group(3) == "상여":
                kind = "bonus"
        elif FNAME_RE[1].match(name):
            m2 = FNAME_RE[1].match(name)
            month = int(m2.group("m_only"))
            year = int(folder_year.group(1)) if folder_year else None
        elif FNAME_RE[2].match(name):
            m3 = FNAME_RE[2].match(name)
            parts = m3.group("intern_date").split(".")
            year, month = 2000 + int(parts[0]), int(parts[1])
            kind = "intern"
        elif "설상여금명세" in name:
            year = int(folder_year.group(1)) if folder_year else None
            month, kind = 2, "bonus"  # 16년 설(2월) 상여 — 구조 상이 시 리포트로 드러남
        else:
            excluded.append(f"{rel} — 파일명 패턴 불일치(수동 확인)")
            continue
        if not year or not month:
            excluded.append(f"{rel} — 연/월 판별 실패")
            continue
        targets.append({"path": p, "rel": rel, "year": year, "month": month, "kind": kind})
    # 동일 (연,월,종류) 중복 → 나중 파일 제외(연도 폴더 정렬상 앞선 것 유지)
    seen = {}
    uniq = []
    for t in targets:
        key = (t["year"], t["month"], t["kind"])
        if key in seen:
            excluded.append(f"{t['rel']} — 중복({seen[key]})")
            continue
        seen[key] = t["rel"]
        uniq.append(t)
    return uniq, excluded


# ---------------------------------------------------------------- parse

def load_workbook_any(path: Path):
    """일반 xlsx 우선, 실패 시 암호(1234) 복호화."""
    try:
        return openpyxl.load_workbook(path, data_only=True), None
    except Exception as e:
        if "not a zip" not in str(e).lower():
            return None, str(e)
    try:
        import msoffcrypto
        with open(path, "rb") as f:
            of = msoffcrypto.OfficeFile(f)
            of.load_key(password=XLSX_PASSWORD)
            buf = io.BytesIO()
            of.decrypt(buf)
        buf.seek(0)
        return openpyxl.load_workbook(buf, data_only=True), None
    except Exception as e2:
        return None, f"암호 해제 실패: {e2}"


def parse_file(t: dict, item_dict: "ItemDict") -> dict:
    """한 파일 → {ok, title, entries[], issues[], header_items[]}. 3단 구조 파싱(§1-1)."""
    wb, err = load_workbook_any(t["path"])
    if err:
        return {"ok": False, "error": err}
    ws = wb.worksheets[0]
    grid = [list(row) for row in ws.iter_rows(values_only=True)]
    nrow = len(grid)

    # 헤더행: '사원번호' 셀이 있는 행
    hdr = next((i for i, row in enumerate(grid)
                if any(squash(v) == "사원번호" for v in row if v is not None)), None)
    if hdr is None:
        return {"ok": False, "error": "헤더('사원번호') 미발견 — 양식 상이"}

    # 그룹행에서 공제 구역 시작 열 (신규 항목 kind 추정용)
    ded_start = None
    for gi in range(max(0, hdr - 2), hdr):
        for c, v in enumerate(grid[gi]):
            if v is not None and "공제" in squash(v):
                ded_start = c
                break
        if ded_start is not None:
            break

    # 3단 헤더맵: (offset, col) → 정규화 항목명 / 합계·인적·정보 열 분리
    header: dict[tuple[int, int], str] = {}
    totals_pos: dict[str, tuple[int, int]] = {}
    person_cols: dict[str, tuple[int, int]] = {}
    skip_pos: set[tuple[int, int]] = set()      # 정보성 열(비과세계·비고 등) — 값 무시
    info_cols: set[int] = set()                 # 정보성 열의 열 번호(세로 병합 대비)
    for k in range(3):
        if hdr + k >= nrow:
            break
        for c, v in enumerate(grid[hdr + k]):
            label = norm_item(v)
            if not label:
                continue
            sq = squash(label)
            if sq in PERSON_LABELS:
                person_cols[sq] = (k, c)
            elif sq == TOTAL_PAY:
                totals_pos["pay"] = (k, c)
            elif sq == TOTAL_DED:
                totals_pos["ded"] = (k, c)
            elif sq == TOTAL_NET:
                totals_pos["net"] = (k, c)
            elif sq in {squash(x) for x in SKIP_INFO}:
                skip_pos.add((k, c))
                if sq != squash("비과세계"):
                    info_cols.add(c)            # 비고·영수인은 열 전체가 정보성
            else:
                header[(k, c)] = label
    totals_set = set(totals_pos.values())

    issues = []
    if len(totals_pos) < 3:
        issues.append(f"합계 열 미발견: {sorted(set(['pay','ded','net']) - set(totals_pos))}")

    title = next((norm_item(v) for row in grid[:4] for v in row
                  if v is not None and "분" in str(v)), None)

    # 직원 3행 블록 순회
    entries = []
    r = hdr + 3
    order = 0
    while r + 2 < nrow + 1 and r < nrow:
        block = [grid[r + k] if r + k < nrow else [] for k in range(3)]
        get = lambda k, c: block[k][c] if k < len(block) and c < len(block[k]) else None
        name_pos = person_cols.get("성명", (0, 1))
        name = norm_item(get(*name_pos))
        empno = get(*person_cols.get("사원번호", (0, 0)))
        # 종료/합계행 판정
        if squash(name) in ("합계", "총계") or (not name and empno is None):
            has_any = any(v is not None and str(v).strip() for row in block for v in row)
            if not has_any or squash(name) in ("합계", "총계"):
                break
        if not name:
            r += 3
            continue

        # 값 셀 전체 순회 — 라벨 있는 위치는 항목 매핑, 라벨 없는 위치는 '미상' 항목으로 보존
        # (선지급분공제 라벨 신설(20년10월) 전 무라벨 기입, 명절상여를 빈 칸에 기입한 관행 대응)
        lines: dict[str, float] = {}
        unknown_here = []
        unlabeled_here = []
        person_pos = set(person_cols.values())
        max_col = max((c for (_, c) in list(header) + list(totals_set)), default=0)
        for k in range(3):
            for c in range(2, max_col + 1):     # A·B(인적) 제외
                if (k, c) in person_pos or (k, c) in totals_set or (k, c) in skip_pos or c in info_cols:
                    continue
                v = num(get(k, c))
                if v is None or v == 0:
                    continue
                is_ded = ded_start is not None and c >= ded_start
                label = header.get((k, c))
                if label is None:
                    item = "unlabeled-ded" if is_ded else "unlabeled-pay"
                    unlabeled_here.append(
                        f"{openpyxl.utils.get_column_letter(c + 1)}{k + 1}단 {v:,.0f}")
                elif k == 0 and (t["rel"], label, name) in NAME_OVERRIDES:
                    item = NAME_OVERRIDES[(t["rel"], label, name)]
                elif k == 0 and label in COLUMN_OVERRIDES.get(t["rel"], {}):
                    item = COLUMN_OVERRIDES[t["rel"]][label]
                else:
                    item = item_dict.resolve(label, is_deduction=is_ded)
                    if item is None:
                        unknown_here.append(label)
                        continue
                lines[item] = lines.get(item, 0) + v

        pay_total = num(get(*totals_pos["pay"])) if "pay" in totals_pos else None
        ded_total = num(get(*totals_pos["ded"])) if "ded" in totals_pos else None
        net_pay = num(get(*totals_pos["net"])) if "net" in totals_pos else None
        # 합계 셀 원본 공란(2016년 초기 양식·인턴 대장 등) → 라인합으로 대체하고 note 기록
        totals_filled = pay_total is None or ded_total is None or net_pay is None

        sum_pay = sum(v for i, v in lines.items() if item_dict.kind(i) == "pay")
        sum_ded = sum(v for i, v in lines.items() if item_dict.kind(i) == "deduction")
        if pay_total is None:
            pay_total = sum_pay
        if ded_total is None:
            ded_total = sum_ded
        if net_pay is None:
            net_pay = pay_total - ded_total
        entry = {
            "row_order": order,
            "emp_no": norm_item(empno) or None,
            "name": name,
            "position_name": norm_item(get(*person_cols.get("직급", (1, 1)))) or None,
            "dept_name": norm_item(get(*person_cols.get("부서", (2, 1)))) or None,
            "hired_at": norm_date(get(*person_cols.get("입사일", (1, 0)))),
            "resigned_at": norm_date(get(*person_cols.get("퇴사일", (2, 0)))),
            "pay_total": pay_total, "deduction_total": ded_total, "net_pay": net_pay,
            "lines": lines, "note": None,
        }
        # 3중 합계 검증 (원 단위, 오차 0)
        errs = []
        if pay_total is not None and round(sum_pay) != round(pay_total):
            errs.append(f"지급 {sum_pay:,.0f}≠{pay_total:,.0f}")
        if ded_total is not None and round(sum_ded) != round(ded_total):
            errs.append(f"공제 {sum_ded:,.0f}≠{ded_total:,.0f}")
        if pay_total is not None and ded_total is not None and net_pay is not None \
                and round(pay_total - ded_total) != round(net_pay):
            errs.append(f"차인 {pay_total - ded_total:,.0f}≠{net_pay:,.0f}")
        if unknown_here:
            errs.append(f"미지 항목 {unknown_here}")
        notes = list(errs)
        if totals_filled:
            notes.append("합계 원본 공란 → 라인합으로 대체")
        if unlabeled_here:
            notes.append("미상값(무라벨 셀): " + ", ".join(unlabeled_here))
        if notes:
            entry["note"] = "; ".join(notes)
        if errs:
            issues.append(f"r{r + 1} {name}: {'; '.join(errs)}")
        entries.append(entry)
        order += 1
        r += 3

    return {"ok": True, "title": title, "entries": entries, "issues": issues,
            "header_items": sorted(set(header.values()))}


# ---------------------------------------------------------------- item dict

class ItemDict:
    """payroll_item_defs 시딩(155)과 동일한 사전 + 미지 항목 수집."""

    def __init__(self):
        self.by_name: dict[str, tuple[str, str]] = {}  # squash(명칭/별칭) → (item_id, kind)
        self.kinds: dict[str, str] = {}
        self.unknown: dict[str, str] = {}              # 신규 발견 항목명 → 추정 kind
        for item_id, name, kind, aliases in SEED_ITEMS:
            self.kinds[item_id] = kind
            for n in [name] + aliases:
                self.by_name[squash(n)] = (item_id, kind)

    def resolve(self, label: str, is_deduction: bool) -> Optional[str]:
        hit = self.by_name.get(squash(label))
        if hit:
            return hit[0]
        self.unknown.setdefault(norm_item(label), "deduction" if is_deduction else "pay")
        return None

    def kind(self, item_id: str) -> str:
        return self.kinds[item_id]


SEED_ITEMS = [
    ("base", "기본급", "pay", []), ("bonus", "상여", "pay", []),
    ("prev-unpaid", "전월미지급금", "pay", []), ("meal", "식대", "pay", []),
    ("overtime", "초과근무수당", "pay", []), ("annual-leave", "연차수당", "pay", []),
    ("incentive", "성과급", "pay", []), ("cert", "자격증수당", "pay", ["자격수당"]),
    ("research", "연구수당", "pay", []), ("comm", "통신비", "pay", []),
    ("trip-lodging", "출장숙박수당", "pay", ["숙박수당"]), ("expert", "전문가활동비", "pay", []),
    ("housing", "주거비지원", "pay", []), ("vehicle", "차량유지비", "pay", []),
    ("misc-pay", "기타수당", "pay", []), ("trip", "출장수당", "pay", []),
    ("childcare", "육아수당", "pay", ["양육수당"]), ("region", "지방파견비", "pay", ["지방파견수당"]),
    ("overtime-meal", "초과근무 식대", "pay", ["초과근무식대"]),
    ("expense-settle", "정산 경비", "pay", ["정산경비"]),
    ("key-talent", "핵심인력성과보상금", "pay", ["핵심인력성과보상금_일반"]),
    ("longevity", "장기근속휴가수당", "pay", []),   # 열 밀림 보정으로 드러난 실항목(158)
    ("unlabeled-pay", "미상 지급", "pay", []),      # 무라벨 셀 값 보존용(원본 위치는 entry.note)
    ("unlabeled-ded", "미상 공제", "deduction", []),
    ("duty-allow", "업무수당", "pay", []),
    ("fixed-ot", "고정연장수당", "pay", []), ("fixed-holiday", "고정휴일수당", "pay", []),
    ("fixed-night", "고정야간수당", "pay", []),
    ("nps", "국민연금", "deduction", []), ("nhis", "건강보험", "deduction", []),
    ("ei", "고용보험", "deduction", []), ("ltc", "장기요양보험료", "deduction", ["장기요양보험"]),
    ("income-tax", "소득세", "deduction", []), ("local-tax", "지방소득세", "deduction", []),
    ("farm-tax", "농특세", "deduction", []),
    ("settle-nps", "정산-국민연금", "deduction", ["국민연금정산"]),
    ("settle-nhis", "정산-건강보험", "deduction", ["건강보험정산"]),
    ("settle-ei", "정산-고용보험", "deduction", ["고용보험정산"]),
    ("settle-ltc", "정산-장기요양", "deduction", ["장기요양정산"]),
    ("settle-income", "정산-근로소득세", "deduction", []),
    ("settle-local", "정산-지방소득세", "deduction", []),
    ("yearend-income", "연말정산소득세", "deduction", ["연말정산 소득세"]),
    ("yearend-local", "연말정산지방소득세", "deduction", ["연말정산 지방소득세"]),
    ("yearend-farm", "연말정산농특세", "deduction", ["연말정산 농특세"]),
    ("advance-ded", "선지급분공제", "deduction", ["공제_선지급분"]),
    ("misc-ded", "기타공제", "deduction", ["그외공제"]),
    ("student-loan", "학자금상환공제", "deduction", ["공제_학자금상환", "공제-학자금상환(기타공제)"]),
]


# ---------------------------------------------------------------- load

def load_db(dsn: str, parsed: list[dict], item_dict: ItemDict, report: list[str],
            force: bool = False) -> None:
    """staging 적재 — ledger 단위 멱등(동일 sha256 스킵, 변경 시 entries 재생성). force=파서 개정 시 전체 재적재."""
    import psycopg

    now = datetime.now().astimezone().isoformat()
    with psycopg.connect(dsn) as conn, conn.cursor() as cur:
        # 미지 항목을 사전에 추가 (item_id = 'auto-' + sha 앞 8)
        for label, kind in item_dict.unknown.items():
            item_id = "auto-" + hashlib.sha256(label.encode()).hexdigest()[:8]
            cur.execute(
                """INSERT INTO payroll_item_defs (item_id, name, kind, display_order, note, created_at)
                   VALUES (%s, %s, %s, 900, '파이프라인 자동 추가', %s)
                   ON CONFLICT (name) DO NOTHING""",
                (item_id, label, kind, now))
            item_dict.by_name[squash(label)] = (item_id, kind)
            item_dict.kinds[item_id] = kind
            report.append(f"- 항목 사전 자동 추가: `{label}` ({kind})")

        # 성명 → employee_id (동명이인은 매칭 보류)
        cur.execute("SELECT employee_id, name FROM employee_profiles")
        by_name: dict[str, list[str]] = {}
        for eid, nm in cur.fetchall():
            by_name.setdefault(squash(nm), []).append(eid)

        n_loaded = n_skipped = 0
        for f in parsed:
            t = f["t"]
            ledger_id = f"pled-{t['year']}{t['month']:02d}-{t['kind']}"
            cur.execute("SELECT source_sha256 FROM payroll_ledgers WHERE ledger_id=%s", (ledger_id,))
            row = cur.fetchone()
            if row and row[0] == f["sha256"] and not force:
                n_skipped += 1
                continue
            cur.execute(
                """INSERT INTO payroll_ledgers
                     (ledger_id, pay_year, pay_month, ledger_kind, title, source, source_file, source_sha256, status, created_at)
                   VALUES (%s,%s,%s,%s,%s,'excel',%s,%s,'confirmed',%s)
                   ON CONFLICT (ledger_id) DO UPDATE SET
                     title=EXCLUDED.title, source_file=EXCLUDED.source_file, source_sha256=EXCLUDED.source_sha256""",
                (ledger_id, t["year"], t["month"], t["kind"], f["title"], t["rel"], f["sha256"], now))
            cur.execute("DELETE FROM payroll_entries WHERE ledger_id=%s", (ledger_id,))
            for e in f["entries"]:
                cands = by_name.get(squash(e["name"]), [])
                employee_id = cands[0] if len(cands) == 1 else None
                if len(cands) > 1:
                    report.append(f"- 동명이인 매칭 보류: {e['name']} ({t['rel']})")
                entry_id = f"{ledger_id}-r{e['row_order']:03d}"
                cur.execute(
                    """INSERT INTO payroll_entries
                         (entry_id, ledger_id, employee_id, emp_no, name, dept_name, position_name,
                          hired_at, resigned_at, pay_total, deduction_total, net_pay, row_order, note)
                       VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                    (entry_id, ledger_id, employee_id, e["emp_no"], e["name"], e["dept_name"],
                     e["position_name"], e["hired_at"], e["resigned_at"],
                     e["pay_total"], e["deduction_total"], e["net_pay"], e["row_order"], e["note"]))
                for item_id, amount in e["lines"].items():
                    cur.execute(
                        "INSERT INTO payroll_entry_lines (line_id, entry_id, item_id, amount) VALUES (%s,%s,%s,%s)",
                        (f"{entry_id}-{item_id}", entry_id, item_id, amount))
            n_loaded += 1
        conn.commit()
    report.append(f"\n**적재 완료**: {n_loaded}개 대장 적재, {n_skipped}개 동일 해시 스킵")


# ---------------------------------------------------------------- main

def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--force", action="store_true", help="동일 해시여도 재적재(파서 개정 시)")
    ap.add_argument("--rematch", action="store_true",
                    help="미매칭 entries 를 employee_profiles 와 재대조만 수행(프로필 신규 등록 후)")
    args = ap.parse_args()

    if args.rematch:
        import psycopg
        dsn = os.environ.get("DATABASE_URL")
        if not dsn:
            raise SystemExit("DATABASE_URL 환경변수가 필요합니다.")
        with psycopg.connect(dsn) as conn, conn.cursor() as cur:
            cur.execute("""
                UPDATE payroll_entries e SET employee_id = p.employee_id
                FROM employee_profiles p
                WHERE e.employee_id IS NULL
                  AND regexp_replace(e.name, '\\s', '', 'g') = regexp_replace(p.name, '\\s', '', 'g')
                  AND 1 = (SELECT count(*) FROM employee_profiles p2
                           WHERE regexp_replace(p2.name, '\\s', '', 'g') = regexp_replace(p.name, '\\s', '', 'g'))""")
            print(f"재대조 완료: {cur.rowcount}행 매칭")
            conn.commit()
        return

    item_dict = ItemDict()
    targets, excluded = inventory()
    parsed, failed = [], []
    total_entries = total_issues = 0

    for t in targets:
        res = parse_file(t, item_dict)
        if not res["ok"]:
            failed.append((t["rel"], res["error"]))
            continue
        if not res["entries"]:
            excluded.append(f"{t['rel']} — 데이터 0명(빈 양식), 적재 제외")
            continue
        res["t"] = t
        res["sha256"] = hashlib.sha256(t["path"].read_bytes()).hexdigest()
        parsed.append(res)
        total_entries += len(res["entries"])
        total_issues += len(res["issues"])

    # 리포트 생성
    n_unlabeled = sum(1 for f in parsed for e in f["entries"]
                      if e["note"] and "미상값" in e["note"])
    rep = [
        "# 급여대장 마이그레이션 검증 리포트",
        f"\n생성: {datetime.now().strftime('%Y-%m-%d %H:%M')} · 모드: {'dry-run' if args.dry_run else 'load'}",
        f"\n## 요약\n- 대상 {len(targets)}개 / 파싱 성공 {len(parsed)}개 / 실패 {len(failed)}개",
        f"- 직원 행 {total_entries}건 / 합계 불일치 {total_issues}건 / 미상값(무라벨 셀) 포함 행 {n_unlabeled}건",
        "- 합계 불일치는 원본 엑셀 자체의 수기 합계 오류 검출분 — 원본값 그대로 적재하고 entry.note 에 기록",
        "- 미상값은 '미상 지급/미상 공제' 항목으로 보존 적재(원본 셀 위치는 entry.note)",
    ]
    if failed:
        rep.append("\n## 파싱 실패")
        rep += [f"- {r} — {e}" for r, e in failed]
    if excluded:
        rep.append("\n## 제외 파일")
        rep += [f"- {x}" for x in excluded]
    if item_dict.unknown:
        rep.append("\n## 사전에 없는 신규 항목 (자동 추가 대상)")
        rep += [f"- `{k}` → {v}" for k, v in item_dict.unknown.items()]
    rep.append("\n## 파일별 상세")
    for f in sorted(parsed, key=lambda x: (x["t"]["year"], x["t"]["month"], x["t"]["kind"])):
        t = f["t"]
        mark = "✅" if not f["issues"] else f"⚠ {len(f['issues'])}건"
        rep.append(f"\n### {t['rel']} — {len(f['entries'])}명 {mark}")
        for iss in f["issues"]:
            rep.append(f"- {iss}")

    if not args.dry_run:
        dsn = os.environ.get("DATABASE_URL")
        if not dsn:
            raise SystemExit("DATABASE_URL 환경변수가 필요합니다 (dry-run 은 --dry-run).")
        load_db(dsn, parsed, item_dict, rep, force=args.force)

    REPORT_PATH.write_text("\n".join(rep), encoding="utf-8")
    print(f"리포트 생성: {REPORT_PATH}")
    print(f"대상 {len(targets)} / 성공 {len(parsed)} / 실패 {len(failed)} / 행 {total_entries} / 경고 {total_issues}")


if __name__ == "__main__":
    main()
