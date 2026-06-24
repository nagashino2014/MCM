#!/usr/bin/env python3
"""직원 인적사항 엑셀 → employee_profiles 기본정보 백필 (일회성).

대상 엑셀: V:\...\직원별 학력사항 및 자격증 보유 현황.xlsx (Sheet1, 데이터 5행~)
성명(A)으로 employee_profiles 를 매칭하여, 기본정보 탭의 다음 항목만 갱신한다.

채우는 항목:
  - 입사일(G)            -> hired_at
  - 주소(E)             -> address
  - 전화번호(F)          -> mobile_phone
  - 메일주소(AK)         -> email
  - 구분(AR) 재직/퇴직    -> status (재직->active, 퇴직->inactive)
  - 주민등록번호(D)       -> resident_registration_encrypted/masked, birth_date, gender, nationality_kind

제외(사용자가 직접 입력): 기술등급(엔지니어링)=eng_grade, 회사 직통 번호=company_phone,
  담당업무=job_duties, 전문분야=specialty_field, 환경부 등급(대행업)=env_grade.
미변경(이미 조직도에 설정됨/엑셀에 없음): dept_id, position_id, employee_no, agent_registered_at, memo.
조직(DB)에 이름이 없는 인원(퇴사자 등)은 스킵.

RRN 암호화는 앱(frontend/lib/admin/employee-records.ts)과 동일:
  key = sha256(EMPLOYEE_PII_ENCRYPTION_KEY||NEXTAUTH_SECRET||AUTH_SECRET||default)
  AES-256-GCM, iv(12B), 출력 = b64(iv).b64(tag).b64(ciphertext)

사용법:
  python scripts/import_basic_info_from_xlsx.py            # dry-run (report만)
  python scripts/import_basic_info_from_xlsx.py --commit   # 실제 반영
"""
from __future__ import annotations

import argparse
import base64
import hashlib
import io
import os
import secrets
from datetime import date, datetime
from pathlib import Path
from typing import Optional

import openpyxl
import psycopg
from cryptography.hazmat.primitives.ciphers.aead import AESGCM

XLSX = Path(r"V:\업무자료\경리업무\사원 경력&인적사항관리정보\직원별 학력사항 및 자격증 보유 현황.xlsx")
DATA_START = 5  # 데이터 시작 행

# 열(1-indexed)
C_NAME = 1    # A 성명
C_RRN = 4     # D 주민등록번호
C_ADDR = 5    # E 주소
C_PHONE = 6   # F 전화번호
C_HIRED = 7   # G 입사일
C_EMAIL = 37  # AK 메일주소
C_STATUS = 44 # AR 구분(재직/퇴직)

DB_URL = os.environ.get(
    "DATABASE_URL",
    "postgresql://mcm:uPFu%23c%24I2ABWSV%5BZ%5D%5DNo3zXr0%5Dfq@localhost:15432/mcm",
)


def pii_key() -> bytes:
    source = (
        os.environ.get("EMPLOYEE_PII_ENCRYPTION_KEY")
        or os.environ.get("NEXTAUTH_SECRET")
        or os.environ.get("AUTH_SECRET")
        or "dev-only-secret-change-me-please-32bytes-1234567890"  # frontend/.env.local AUTH_SECRET
    )
    return hashlib.sha256(source.encode("utf-8")).digest()


def encrypt_rrn(digits: str) -> str:
    key = pii_key()
    iv = secrets.token_bytes(12)
    ct_and_tag = AESGCM(key).encrypt(iv, digits.encode("utf-8"), None)
    ct, tag = ct_and_tag[:-16], ct_and_tag[-16:]
    return ".".join(base64.b64encode(b).decode("ascii") for b in (iv, tag, ct))


CODE_MAP = {
    "1": (1900, "male", "domestic"), "2": (1900, "female", "domestic"),
    "3": (2000, "male", "domestic"), "4": (2000, "female", "domestic"),
    "5": (1900, "male", "foreign"), "6": (1900, "female", "foreign"),
    "7": (2000, "male", "foreign"), "8": (2000, "female", "foreign"),
    "9": (1800, "male", "domestic"), "0": (1800, "female", "domestic"),
}


def parse_rrn(value) -> Optional[dict]:
    digits = "".join(ch for ch in str(value or "") if ch.isdigit())
    if len(digits) != 13:
        return None
    yy, mm, dd, g = int(digits[:2]), int(digits[2:4]), int(digits[4:6]), digits[6]
    meta = CODE_MAP.get(g)
    if not meta:
        return None
    century, gender, nat = meta
    year = century + yy
    try:
        d = date(year, mm, dd)
    except ValueError:
        return None
    return {
        "encrypted": encrypt_rrn(digits),
        "masked": f"{digits[:6]}-{g}******",
        "birth": d.isoformat(),
        "gender": gender,
        "nat": nat,
    }


def to_iso_date(v) -> Optional[str]:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    s = str(v).strip()
    return s or None


def clean(v) -> Optional[str]:
    s = str(v).strip() if v is not None else ""
    return s or None


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--commit", action="store_true", help="실제 DB 반영")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["Sheet1"]

    excel_rows = []
    for r in range(DATA_START, ws.max_row + 1):
        name = clean(ws.cell(r, C_NAME).value)
        if not name:
            continue
        status_raw = clean(ws.cell(r, C_STATUS).value) or ""
        excel_rows.append({
            "row": r,
            "name": name,
            "rrn": parse_rrn(ws.cell(r, C_RRN).value),
            "rrn_raw": clean(ws.cell(r, C_RRN).value),
            "addr": clean(ws.cell(r, C_ADDR).value),
            "phone": clean(ws.cell(r, C_PHONE).value),
            "hired": to_iso_date(ws.cell(r, C_HIRED).value),
            "email": clean(ws.cell(r, C_EMAIL).value),
            "status": "inactive" if "퇴직" in status_raw else "active",
            "status_raw": status_raw,
        })

    conn = psycopg.connect(DB_URL, sslmode="require", connect_timeout=12)
    cur = conn.cursor()
    cur.execute("SELECT employee_id, name, status FROM employee_profiles")
    by_name: dict[str, list[tuple]] = {}
    for emp_id, nm, st in cur.fetchall():
        by_name.setdefault(nm, []).append((emp_id, st))

    rpt = io.StringIO()
    rpt.write(f"엑셀 행 수: {len(excel_rows)} / DB 직원 수: {sum(len(v) for v in by_name.values())}\n")
    rpt.write(f"모드: {'COMMIT' if args.commit else 'DRY-RUN'}\n\n")

    updated, skipped, dup, rrn_fail = [], [], [], []
    matched_names = set()
    now = datetime.utcnow().isoformat()

    for er in excel_rows:
        cands = by_name.get(er["name"])
        if not cands:
            skipped.append(er)
            continue
        if len(cands) > 1:
            dup.append((er["name"], len(cands)))
        emp_id = cands[0][0]
        matched_names.add(er["name"])
        if er["rrn_raw"] and not er["rrn"]:
            rrn_fail.append((er["name"], er["rrn_raw"]))

        sets, vals = [], []
        def put(col, val):
            sets.append(f"{col} = %s")
            vals.append(val)

        put("hired_at", er["hired"])
        put("address", er["addr"])
        put("mobile_phone", er["phone"])
        put("email", er["email"])
        # status(재직상태)는 갱신하지 않음: 매칭된 32명은 모두 현 조직 active이고,
        # 엑셀 AR의 '퇴직' 표기가 일부 stale(예: 권오종, 최근 입사)이라 inactive 설정 시
        # 조직도에서 사라질 위험이 있어 보존한다. (퇴직 표기는 리포트로만 안내)
        if er["rrn"]:
            put("resident_registration_encrypted", er["rrn"]["encrypted"])
            put("resident_registration_masked", er["rrn"]["masked"])
            put("birth_date", er["rrn"]["birth"])
            put("gender", er["rrn"]["gender"])
            put("nationality_kind", er["rrn"]["nat"])
        put("updated_at", now)
        vals.append(emp_id)

        if args.commit:
            cur.execute(
                f"UPDATE employee_profiles SET {', '.join(sets)} WHERE employee_id = %s",
                vals,
            )
        updated.append(er)

    db_not_in_excel = sorted(set(by_name) - matched_names)

    rpt.write(f"== 갱신 대상 (매칭됨): {len(updated)}명 ==\n")
    for er in updated:
        rpt.write(
            f"  {er['name']:6s} | 입사 {er['hired'] or '-'} | {er['phone'] or '-'} "
            f"| {er['email'] or '-'} | RRN {'O' if er['rrn'] else 'X'} | {er['status']} "
            f"| 주소 {(er['addr'] or '')[:20]}\n"
        )
    rpt.write(f"\n== 엑셀에 있으나 DB(조직)에 없음 → 스킵: {len(skipped)}명 ==\n")
    rpt.write("  " + ", ".join(e["name"] for e in skipped) + "\n")
    rpt.write(f"\n== DB에 있으나 엑셀에 없음(매칭 실패): {len(db_not_in_excel)}명 ==\n")
    rpt.write("  " + ", ".join(db_not_in_excel) + "\n")
    if dup:
        rpt.write(f"\n== 동명이인(DB 다중) — 첫 행만 갱신: ==\n  {dup}\n")
    if rrn_fail:
        rpt.write(f"\n== RRN 파싱 실패: ==\n  {rrn_fail}\n")
    inact = [e["name"] for e in updated if e["status"] == "inactive"]
    if inact:
        rpt.write(f"\n== status=inactive(퇴직)로 설정됨: ==\n  {inact}\n")

    if args.commit:
        conn.commit()
        rpt.write("\n>>> COMMIT 완료\n")
    else:
        conn.rollback()
        rpt.write("\n>>> DRY-RUN (롤백, 변경 없음)\n")
    conn.close()

    Path(r"C:\CodingProject\MCM\import_basic_info_report.txt").write_text(rpt.getvalue(), encoding="utf-8")
    print("report ->", r"C:\CodingProject\MCM\import_basic_info_report.txt")


if __name__ == "__main__":
    main()
