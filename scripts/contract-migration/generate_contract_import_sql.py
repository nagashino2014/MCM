#!/usr/bin/env python3
"""
Generate idempotent SQL to import the legacy contract workbook.

The generated SQL is intended to be applied after infra/aws/004_contract_management.sql.
It imports:
- legal_entities from 업체정보 and contract counterparties
- contracts from 계약입력
- contract_payment_milestones from 계약입력 stage columns
- contract_change_events for changed contracts
- contract_import_batches summary metadata

It does not import tax invoice PDF binaries. Those are registered later through
the web upload flow.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


DEFAULT_WORKBOOK = r"v:\계약\계약현황.xlsm"

SHEET_CONTRACT_INPUT = "계약입력"
SHEET_COMPANY = "업체정보"

CONTRACT_HEADERS = {
    "row_no": 0,
    "contract_title": 1,
    "counterparty_name": 2,
    "service_type": 3,
    "order_amount": 4,
    "contract_date": 5,
    "started_at": 6,
    "ended_at": 7,
    "industry_category": 8,
    "legacy_company_id": 9,
    "collection_progress": 10,
    "quote_issued": 11,
    "quote_no": 12,
    "quote_issued_at": 13,
    "change_flag": 14,
    "change_date": 15,
    "changed_company_name_1": 16,
    "changed_contract_title_1": 17,
    "changed_company_name_2": 18,
    "changed_contract_title_2": 19,
    "changed_at": 20,
    "entered_at": 21,
    "changed_amount": 22,
    "changed_service_period": 23,
    "changed_payment_terms": 24,
    "changed_misc": 25,
    "previous_amount": 26,
    "delta_amount": 27,
    "change_detail": 28,
}

MILESTONE_COLUMNS = [
    ("advance", "선급금", 1, 29, 30, 31, 32, 33, 34, 35, 36, 37),
    ("mid1", "중도금1", 2, 38, 39, 40, 41, 42, 43, 44, 45, 46),
    ("mid2", "중도금2", 3, 47, 48, 49, 50, 51, 52, 53, 54, 55),
    ("mid3", "중도금3", 4, 56, 57, 58, 59, 60, 61, 62, 63, 64),
    ("completion", "준공금", 5, 65, 66, 67, 68, 69, 70, 71, 72, 73),
]

COMPANY_HEADERS = {
    "company_name": 1,
    "headquarter_id": 2,
    "site_id_1": 3,
    "site_id_2": 4,
    "site_id_3": 5,
    "business_registration_no": 6,
    "corporate_registration_no": 7,
    "representative_name": 8,
    "headquarter_address": 9,
    "contract_entity_type": 10,
    "site_name_1": 15,
    "site_name_2": 16,
    "site_name_3": 17,
    "site_business_no_1": 18,
    "site_business_no_2": 19,
    "site_business_no_3": 20,
    "site_address_1": 21,
    "site_address_2": 22,
    "site_address_3": 23,
    "integrated_permit_target": 24,
}


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def normalize_company_name(value: Any) -> str:
    text = normalize_text(value)
    text = re.sub(r"\(주\)|㈜|주식회사", "", text, flags=re.IGNORECASE)
    text = re.sub(r"[\s·ㆍ\.\-_\(\)（）]", "", text)
    return text.lower()


def normalize_digits(value: Any) -> str | None:
    digits = re.sub(r"\D+", "", normalize_text(value))
    return digits or None


def as_iso_date(value: Any) -> str | None:
    if isinstance(value, datetime):
        if value.year <= 1900:
            return None
        return value.date().isoformat()
    if isinstance(value, date):
        if value.year <= 1900:
            return None
        return value.isoformat()
    text = normalize_text(value)
    if not text or text in {"0", "00:00:00"}:
        return None
    return text


def number(value: Any) -> float | None:
    if isinstance(value, (int, float)):
        return float(value)
    text = normalize_text(value).replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def boolish(value: Any) -> int:
    text = normalize_text(value)
    if not text:
        return 0
    return 0 if text in {"0", "False", "false", "N", "n", "X", "x"} else 1


def get(row: tuple[Any, ...], index: int) -> Any:
    return row[index] if index < len(row) else None


def digest_id(prefix: str, *parts: Any, length: int = 16) -> str:
    raw = "|".join(normalize_text(part) for part in parts)
    return prefix + hashlib.sha256(raw.encode("utf-8")).hexdigest()[:length]


def sql_value(value: Any) -> str:
    if value is None:
        return "NULL"
    if isinstance(value, bool):
        return "1" if value else "0"
    if isinstance(value, (int, float)):
        return str(value)
    text = normalize_text(value)
    if text == "":
        return "NULL"
    return "'" + text.replace("'", "''") + "'"


def sql_json(value: Any) -> str:
    return "'" + json.dumps(value, ensure_ascii=False).replace("'", "''") + "'::jsonb"


def row_dict(row: tuple[Any, ...], headers: dict[str, int]) -> dict[str, Any]:
    return {key: get(row, idx) for key, idx in headers.items()}


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate legacy contract import SQL.")
    parser.add_argument("--workbook", default=DEFAULT_WORKBOOK)
    parser.add_argument("--out", required=True)
    parser.add_argument("--batch-id")
    args = parser.parse_args()

    workbook_path = Path(args.workbook)
    if not workbook_path.exists():
        raise SystemExit(f"Workbook not found: {workbook_path}")

    batch_id = args.batch_id or "cimp_" + datetime.now().strftime("%Y%m%d%H%M%S")
    now = datetime.now().isoformat(timespec="seconds")
    wb = load_workbook(workbook_path, read_only=True, keep_vba=True, data_only=True)
    contract_ws = wb[SHEET_CONTRACT_INPUT]
    company_ws = wb[SHEET_COMPANY]

    company_by_id: dict[str, dict[str, Any]] = {}
    entity_by_key: dict[str, dict[str, Any]] = {}

    for source_row, row in enumerate(company_ws.iter_rows(min_row=5, values_only=True), start=5):
      item = row_dict(row, COMPANY_HEADERS)
      company_name = normalize_text(item["company_name"])
      if not company_name and not normalize_text(item["headquarter_id"]):
          continue
      ids = [
          normalize_text(item["headquarter_id"]),
          normalize_text(item["site_id_1"]),
          normalize_text(item["site_id_2"]),
          normalize_text(item["site_id_3"]),
      ]
      for legacy_id in [v for v in ids if v and v != "0"]:
          company_by_id[legacy_id] = {**item, "source_row": source_row, "legacy_company_id": legacy_id}
      entity_key = normalize_text(item["headquarter_id"]) or normalize_company_name(company_name)
      if entity_key:
          entity_by_key[entity_key] = {**item, "source_row": source_row, "legacy_company_id": normalize_text(item["headquarter_id"]) or None}

    contracts: list[dict[str, Any]] = []
    for source_row, row in enumerate(contract_ws.iter_rows(min_row=2, values_only=True), start=2):
        contract_title = normalize_text(get(row, CONTRACT_HEADERS["contract_title"]))
        if not contract_title:
            continue
        item = row_dict(row, CONTRACT_HEADERS)
        item["_row"] = row
        item["source_row"] = source_row
        legacy_company_id = normalize_text(item["legacy_company_id"])
        counterparty_name = normalize_text(item["counterparty_name"])
        if legacy_company_id and legacy_company_id not in company_by_id:
            company_by_id[legacy_company_id] = {
                "company_name": counterparty_name,
                "headquarter_id": legacy_company_id,
                "legacy_company_id": legacy_company_id,
                "source_row": None,
            }
        if legacy_company_id:
            entity_by_key.setdefault(
                legacy_company_id,
                {
                    "company_name": counterparty_name,
                    "headquarter_id": legacy_company_id,
                    "legacy_company_id": legacy_company_id,
                    "source_row": None,
                },
            )
        contracts.append(item)

    lines: list[str] = [
        "BEGIN;",
        "SET CONSTRAINTS ALL DEFERRED;",
        "",
        "INSERT INTO contract_import_batches (batch_id, source_workbook, workbook_size_bytes, workbook_modified_at, status, summary_json, created_at, updated_at)",
        "VALUES ("
        + ", ".join(
            [
                sql_value(batch_id),
                sql_value(str(workbook_path)),
                sql_value(workbook_path.stat().st_size),
                sql_value(datetime.fromtimestamp(workbook_path.stat().st_mtime).isoformat()),
                sql_value("applied"),
                sql_json({"contracts": len(contracts), "entities": len(entity_by_key), "source": "legacy_contract_workbook"}),
                sql_value(now),
                sql_value(now),
            ]
        )
        + ") ON CONFLICT (batch_id) DO UPDATE SET summary_json = EXCLUDED.summary_json, updated_at = EXCLUDED.updated_at;",
        "",
    ]

    for key, item in sorted(entity_by_key.items()):
        company_name = normalize_text(item.get("company_name")) or key
        entity_id = digest_id("ent_", normalize_text(item.get("legacy_company_id")) or normalize_company_name(company_name))
        lines.append(
            "INSERT INTO legal_entities "
            "(entity_id, entity_name, business_registration_no, address, memo, created_at, updated_at, "
            "corporate_registration_no, representative_name, contract_entity_type, legacy_company_id, source_workbook, source_sheet, source_row) VALUES ("
            + ", ".join(
                [
                    sql_value(entity_id),
                    sql_value(company_name),
                    sql_value(normalize_digits(item.get("business_registration_no"))),
                    sql_value(item.get("headquarter_address")),
                    sql_value("계약현황 엑셀 이관"),
                    sql_value(now),
                    sql_value(now),
                    sql_value(normalize_digits(item.get("corporate_registration_no"))),
                    sql_value(item.get("representative_name")),
                    sql_value(item.get("contract_entity_type")),
                    sql_value(item.get("legacy_company_id")),
                    sql_value(str(workbook_path)),
                    sql_value(SHEET_COMPANY),
                    sql_value(item.get("source_row")),
                ]
            )
            + ") ON CONFLICT (entity_id) DO UPDATE SET "
            + "entity_name = EXCLUDED.entity_name, business_registration_no = COALESCE(EXCLUDED.business_registration_no, legal_entities.business_registration_no), "
            + "address = COALESCE(EXCLUDED.address, legal_entities.address), corporate_registration_no = COALESCE(EXCLUDED.corporate_registration_no, legal_entities.corporate_registration_no), "
            + "representative_name = COALESCE(EXCLUDED.representative_name, legal_entities.representative_name), contract_entity_type = COALESCE(EXCLUDED.contract_entity_type, legal_entities.contract_entity_type), "
            + "legacy_company_id = COALESCE(EXCLUDED.legacy_company_id, legal_entities.legacy_company_id), updated_at = EXCLUDED.updated_at;"
        )

    lines.append("")

    for item in contracts:
        row = item["_row"]
        legacy_company_id = normalize_text(item["legacy_company_id"])
        counterparty_name = normalize_text(item["counterparty_name"])
        entity_seed = legacy_company_id or normalize_company_name(counterparty_name)
        entity_id = digest_id("ent_", entity_seed)
        contract_date = as_iso_date(item["contract_date"])
        order_amount = number(item["order_amount"])
        # Preserve every legacy workbook row. Some rows intentionally share the
        # same company/title/date/amount and should be reviewed by the dedupe UI
        # later, not merged during import.
        contract_id = digest_id("ctr_", SHEET_CONTRACT_INPUT, item["source_row"], legacy_company_id, item["contract_title"], contract_date, order_amount)
        current_amount = number(item["changed_amount"]) or order_amount
        lines.append(
            "INSERT INTO contracts "
            "(contract_id, facility_id, counterparty_entity_id, contract_title, service_type, contract_status, contract_amount, "
            "started_at, ended_at, memo, created_at, updated_at, legacy_contract_no, legacy_company_id, contract_direction, "
            "industry_category, contract_date, original_amount, current_amount, source_workbook, source_sheet, source_row, migration_batch_id) VALUES ("
            + ", ".join(
                [
                    sql_value(contract_id),
                    "NULL",
                    sql_value(entity_id),
                    sql_value(item["contract_title"]),
                    sql_value(item["service_type"]),
                    sql_value("active"),
                    sql_value(order_amount),
                    sql_value(as_iso_date(item["started_at"])),
                    sql_value(as_iso_date(item["ended_at"])),
                    sql_value("계약현황 엑셀 이관"),
                    sql_value(now),
                    sql_value(now),
                    sql_value(item["row_no"]),
                    sql_value(legacy_company_id),
                    sql_value("sales"),
                    sql_value(item["industry_category"]),
                    sql_value(contract_date),
                    sql_value(order_amount),
                    sql_value(current_amount),
                    sql_value(str(workbook_path)),
                    sql_value(SHEET_CONTRACT_INPUT),
                    sql_value(item["source_row"]),
                    sql_value(batch_id),
                ]
            )
            + ") ON CONFLICT (contract_id) DO UPDATE SET "
            + "contract_title = EXCLUDED.contract_title, service_type = EXCLUDED.service_type, contract_amount = EXCLUDED.contract_amount, "
            + "started_at = EXCLUDED.started_at, ended_at = EXCLUDED.ended_at, current_amount = EXCLUDED.current_amount, updated_at = EXCLUDED.updated_at;"
        )

        for stage_key, stage_label, stage_order, amount_idx, ratio_idx, collect_ratio_idx, collected_idx, collected_amount_idx, collected_at_idx, issued_idx, issued_at_idx, terms_idx in MILESTONE_COLUMNS:
            amount = number(get(row, amount_idx))
            issued = boolish(get(row, issued_idx))
            collected = boolish(get(row, collected_idx))
            if not amount and not issued and not collected:
                continue
            milestone_id = digest_id("mil_", contract_id, stage_key)
            lines.append(
                "INSERT INTO contract_payment_milestones "
                "(milestone_id, contract_id, stage_key, stage_label, stage_order, amount, amount_ratio, collection_ratio, "
                "invoice_issued, invoice_issued_at, invoice_amount, payment_collected, payment_collected_at, collected_amount, "
                "payment_terms, legacy_source_sheet, legacy_source_row, created_at, updated_at) VALUES ("
                + ", ".join(
                    [
                        sql_value(milestone_id),
                        sql_value(contract_id),
                        sql_value(stage_key),
                        sql_value(stage_label),
                        sql_value(stage_order),
                        sql_value(amount),
                        sql_value(number(get(row, ratio_idx))),
                        sql_value(number(get(row, collect_ratio_idx))),
                        sql_value(issued),
                        sql_value(as_iso_date(get(row, issued_at_idx))),
                        sql_value(amount if issued else None),
                        sql_value(collected),
                        sql_value(as_iso_date(get(row, collected_at_idx))),
                        sql_value(number(get(row, collected_amount_idx)) or (amount if collected else None)),
                        sql_value(get(row, terms_idx)),
                        sql_value(SHEET_CONTRACT_INPUT),
                        sql_value(item["source_row"]),
                        sql_value(now),
                        sql_value(now),
                    ]
                )
                + ") ON CONFLICT (contract_id, stage_key) DO UPDATE SET "
                + "amount = EXCLUDED.amount, amount_ratio = EXCLUDED.amount_ratio, collection_ratio = EXCLUDED.collection_ratio, "
                + "invoice_issued = EXCLUDED.invoice_issued, invoice_issued_at = EXCLUDED.invoice_issued_at, invoice_amount = EXCLUDED.invoice_amount, "
                + "payment_collected = EXCLUDED.payment_collected, payment_collected_at = EXCLUDED.payment_collected_at, collected_amount = EXCLUDED.collected_amount, "
                + "payment_terms = EXCLUDED.payment_terms, updated_at = EXCLUDED.updated_at;"
            )

        if boolish(item["change_flag"]) or number(item["delta_amount"]):
            change_id = digest_id("chg_", contract_id, item["change_date"], item["delta_amount"], item["change_detail"])
            lines.append(
                "INSERT INTO contract_change_events "
                "(change_id, contract_id, changed_at, previous_amount, delta_amount, changed_service_period, changed_payment_terms, detail, source_sheet, source_row, created_at, updated_at) VALUES ("
                + ", ".join(
                    [
                        sql_value(change_id),
                        sql_value(contract_id),
                        sql_value(as_iso_date(item["change_date"]) or as_iso_date(item["changed_at"])),
                        sql_value(number(item["previous_amount"])),
                        sql_value(number(item["delta_amount"])),
                        sql_value(item["changed_service_period"]),
                        sql_value(item["changed_payment_terms"]),
                        sql_value(item["change_detail"]),
                        sql_value(SHEET_CONTRACT_INPUT),
                        sql_value(item["source_row"]),
                        sql_value(now),
                        sql_value(now),
                    ]
                )
                + ") ON CONFLICT (change_id) DO UPDATE SET detail = EXCLUDED.detail, updated_at = EXCLUDED.updated_at;"
            )

    lines.extend(
        [
            "",
            "COMMIT;",
            "",
            "SELECT COUNT(*) AS contracts FROM contracts;",
            "SELECT COUNT(*) AS milestones FROM contract_payment_milestones;",
            "SELECT COUNT(*) AS contract_entities FROM legal_entities WHERE source_sheet = '업체정보' OR memo = '계약현황 엑셀 이관';",
        ]
    )

    out_path = Path(args.out)
    out_path.write_text("\n".join(lines), encoding="utf-8")
    print(json.dumps({"out": str(out_path), "batchId": batch_id, "contracts": len(contracts), "entities": len(entity_by_key)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
