#!/usr/bin/env python3
"""
Dry-run analyzer for the legacy contract workbook.

This script intentionally does not execute VBA macros and does not write to the
application database. It reads cached workbook values, summarizes the migration
surface, and emits issue candidates that should be reviewed before importing.
"""

from __future__ import annotations

import argparse
import json
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime, time
from hashlib import sha256
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


DEFAULT_WORKBOOK = r"v:\계약\계약현황.xlsm"


SHEET_CONTRACT_INPUT = "계약입력"
SHEET_BILLING_LEDGER = "Sheet4"
SHEET_COMPANY = "업체정보"
SHEET_CONTACT = "업체담당자정보현황"
SHEET_FACILITY_DETAIL = "통합대상사업장시설현황"


CONTRACT_HEADERS = {
    "row_no": 0,
    "contract_title": 1,
    "counterparty_name": 2,
    "service_type": 3,
    "order_amount": 4,
    "contract_date": 5,
    "started_at": 6,
    "ended_at": 7,
    "industry_category": 8,
    "legacy_company_id": 9,
    "collection_progress": 10,
    "quote_issued": 11,
    "quote_no": 12,
    "quote_issued_at": 13,
    "change_flag": 14,
    "change_date": 15,
    "changed_company_name_1": 16,
    "changed_contract_title_1": 17,
    "changed_company_name_2": 18,
    "changed_contract_title_2": 19,
    "changed_at": 20,
    "entered_at": 21,
    "changed_amount": 22,
    "changed_service_period": 23,
    "changed_payment_terms": 24,
    "changed_misc": 25,
    "previous_amount": 26,
    "delta_amount": 27,
    "change_detail": 28,
}


MILESTONE_COLUMNS = [
    ("advance", "선급금", 29, 30, 31, 32, 33, 34, 35, 36, 37),
    ("mid1", "중도금1", 38, 39, 40, 41, 42, 43, 44, 45, 46),
    ("mid2", "중도금2", 47, 48, 49, 50, 51, 52, 53, 54, 55),
    ("mid3", "중도금3", 56, 57, 58, 59, 60, 61, 62, 63, 64),
    ("completion", "준공금", 65, 66, 67, 68, 69, 70, 71, 72, 73),
]


COMPANY_HEADERS = {
    "company_name": 1,
    "headquarter_id": 2,
    "site_id_1": 3,
    "site_id_2": 4,
    "site_id_3": 5,
    "business_registration_no": 6,
    "corporate_registration_no": 7,
    "representative_name": 8,
    "headquarter_address": 9,
    "contract_entity_type": 10,
    "main_industry_code": 11,
    "site_industry_code_1": 12,
    "site_industry_code_2": 13,
    "site_industry_code_3": 14,
    "site_name_1": 15,
    "site_name_2": 16,
    "site_name_3": 17,
    "site_business_no_1": 18,
    "site_business_no_2": 19,
    "site_business_no_3": 20,
    "site_address_1": 21,
    "site_address_2": 22,
    "site_address_3": 23,
    "integrated_permit_target": 24,
    "entered_at": 25,
    "updated_at": 26,
}


@dataclass
class SheetStats:
    rows: int
    cols: int
    nonempty: int


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return ""
    text = str(value).strip()
    return re.sub(r"\s+", " ", text)


def normalize_company_name(value: Any) -> str:
    text = normalize_text(value)
    text = re.sub(r"\(주\)|㈜|주식회사", "", text, flags=re.IGNORECASE)
    text = re.sub(r"[\s·ㆍ\.\-_\(\)（）]", "", text)
    return text.lower()


def normalize_digits(value: Any) -> str:
    return re.sub(r"\D+", "", normalize_text(value))


def number(value: Any) -> float:
    if isinstance(value, (int, float)):
        return float(value)
    text = normalize_text(value).replace(",", "")
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def boolish(value: Any) -> bool:
    text = normalize_text(value)
    if not text:
        return False
    return text not in {"0", "False", "false", "N", "n", "X", "x"}


def as_iso_date(value: Any) -> str | None:
    if isinstance(value, datetime):
        if value.year <= 1900:
            return None
        return value.date().isoformat()
    if isinstance(value, date):
        if value.year <= 1900:
            return None
        return value.isoformat()
    text = normalize_text(value)
    if not text or text in {"0", "00:00:00"}:
        return None
    return text


def fiscal_quarter(value: Any) -> tuple[int | None, int | None]:
    dt: date | None = None
    if isinstance(value, datetime):
        dt = value.date()
    elif isinstance(value, date):
        dt = value
    if not dt or dt.year <= 1900:
        return None, None
    return dt.year, (dt.month - 1) // 3 + 1


def get(row: tuple[Any, ...], index: int) -> Any:
    return row[index] if index < len(row) else None


def row_to_contract(row: tuple[Any, ...], source_row: int) -> dict[str, Any]:
    item = {key: get(row, idx) for key, idx in CONTRACT_HEADERS.items()}
    item["source_row"] = source_row
    item["legacy_key"] = make_contract_legacy_key(item)
    return item


def make_contract_legacy_key(item: dict[str, Any]) -> str:
    parts = [
        normalize_text(item.get("legacy_company_id")),
        normalize_text(item.get("contract_title")),
        as_iso_date(item.get("contract_date")) or "",
        str(int(number(item.get("order_amount")))),
    ]
    return sha256("|".join(parts).encode("utf-8")).hexdigest()[:24]


def iter_data_rows(ws: Any, min_row: int) -> Iterable[tuple[int, tuple[Any, ...]]]:
    for index, row in enumerate(ws.iter_rows(min_row=min_row, values_only=True), start=min_row):
        yield index, row


def count_nonempty(ws: Any) -> int:
    total = 0
    for row in ws.iter_rows(values_only=True):
        total += sum(1 for value in row if value is not None)
    return total


def analyze_contracts(ws: Any) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, Any]]:
    contracts: list[dict[str, Any]] = []
    issues: list[dict[str, Any]] = []
    duplicate_keys: defaultdict[str, list[int]] = defaultdict(list)

    for source_row, row in iter_data_rows(ws, 2):
        if not normalize_text(get(row, CONTRACT_HEADERS["contract_title"])):
            continue
        item = row_to_contract(row, source_row)
        contracts.append(item)
        duplicate_keys[item["legacy_key"]].append(source_row)

        amount = number(item.get("order_amount"))
        if amount <= 0:
            issues.append(issue("contract_amount", source_row, "수주총액이 비어 있거나 0 이하입니다.", item))
        if not normalize_text(item.get("legacy_company_id")):
            issues.append(issue("missing_company_id", source_row, "업체ID가 비어 있습니다.", item))
        if not as_iso_date(item.get("contract_date")):
            issues.append(issue("missing_contract_date", source_row, "계약일이 비어 있거나 날짜로 인식되지 않습니다.", item))
        if not normalize_text(item.get("counterparty_name")):
            issues.append(issue("missing_counterparty", source_row, "계약상대자가 비어 있습니다.", item))

    for key, rows in duplicate_keys.items():
        if len(rows) > 1:
            issues.append(
                {
                    "type": "duplicate_contract_candidate",
                    "sourceRows": rows,
                    "legacyKey": key,
                    "message": "업체ID/계약명/계약일/금액 조합이 중복됩니다.",
                }
            )

    summary = {
        "count": len(contracts),
        "uniqueCounterparties": len({normalize_company_name(c.get("counterparty_name")) for c in contracts if c.get("counterparty_name")}),
        "uniqueCompanyIds": len({normalize_text(c.get("legacy_company_id")) for c in contracts if c.get("legacy_company_id")}),
        "totalOrderAmount": round(sum(number(c.get("order_amount")) for c in contracts)),
        "serviceTypes": Counter(normalize_text(c.get("service_type")) for c in contracts if c.get("service_type")).most_common(),
        "contractYears": Counter((as_iso_date(c.get("contract_date")) or "")[:4] for c in contracts if as_iso_date(c.get("contract_date"))).most_common(),
    }
    return contracts, issues, summary


def analyze_milestones(contracts: list[dict[str, Any]]) -> dict[str, Any]:
    totals: dict[str, dict[str, Any]] = {}
    grand_total = 0.0
    for stage_key, stage_label, amount_idx, ratio_idx, collect_ratio_idx, collected_idx, collected_amount_idx, collected_at_idx, issued_idx, issued_at_idx, terms_idx in MILESTONE_COLUMNS:
        stage_total = 0.0
        stage_count = 0
        issued_count = 0
        collected_count = 0
        for item in contracts:
            row = item["_row"] if "_row" in item else None
            if row is None:
                continue
            amount = number(get(row, amount_idx))
            if amount > 0:
                stage_count += 1
                stage_total += amount
                grand_total += amount
            if boolish(get(row, issued_idx)):
                issued_count += 1
            if boolish(get(row, collected_idx)):
                collected_count += 1
        totals[stage_key] = {
            "label": stage_label,
            "count": stage_count,
            "amount": round(stage_total),
            "issuedCount": issued_count,
            "collectedCount": collected_count,
            "columnIndexes": {
                "amount": amount_idx,
                "ratio": ratio_idx,
                "collectionRatio": collect_ratio_idx,
                "collected": collected_idx,
                "collectedAmount": collected_amount_idx,
                "collectedAt": collected_at_idx,
                "issued": issued_idx,
                "issuedAt": issued_at_idx,
                "terms": terms_idx,
            },
        }
    return {"totalMilestoneAmount": round(grand_total), "stages": totals}


def analyze_billing_ledger(ws: Any) -> dict[str, Any]:
    rows = []
    for source_row, row in iter_data_rows(ws, 2):
        # Sheet4 contains long formula-filled ranges. Treat a row as meaningful
        # only when it has a positive amount or an explicit issued/collected flag.
        if number(get(row, 6)) > 0 or boolish(get(row, 8)) or boolish(get(row, 9)):
            rows.append(row)
    return {
        "count": len(rows),
        "amount": round(sum(number(get(row, 6)) for row in rows)),
        "types": Counter(normalize_text(get(row, 5)) for row in rows if get(row, 5)).most_common(),
        "issued": Counter(normalize_text(get(row, 9)) for row in rows).most_common(),
        "collected": Counter(normalize_text(get(row, 8)) for row in rows).most_common(),
    }


def analyze_companies(ws: Any) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, Any]]:
    companies: list[dict[str, Any]] = []
    issues: list[dict[str, Any]] = []
    seen_ids: defaultdict[str, list[int]] = defaultdict(list)
    seen_brn: defaultdict[str, list[int]] = defaultdict(list)

    for source_row, row in iter_data_rows(ws, 5):
        if not normalize_text(get(row, COMPANY_HEADERS["company_name"])) and not normalize_text(get(row, COMPANY_HEADERS["headquarter_id"])):
            continue
        item = {key: get(row, idx) for key, idx in COMPANY_HEADERS.items()}
        item["source_row"] = source_row
        companies.append(item)

        for key in ("headquarter_id", "site_id_1", "site_id_2", "site_id_3"):
            value = normalize_text(item.get(key))
            if value and value != "0":
                seen_ids[value].append(source_row)
        brn = normalize_digits(item.get("business_registration_no"))
        if brn:
            seen_brn[brn].append(source_row)
            if len(brn) != 10:
                issues.append(issue("invalid_business_registration_no", source_row, "사업자등록번호가 10자리가 아닙니다.", item))
        corporate_no = normalize_digits(item.get("corporate_registration_no"))
        if corporate_no and len(corporate_no) != 13:
            issues.append(issue("invalid_corporate_registration_no", source_row, "법인번호가 13자리가 아닙니다.", item))

    for company_id, rows in seen_ids.items():
        if len(rows) > 1:
            issues.append({"type": "duplicate_company_id", "sourceRows": rows, "legacyCompanyId": company_id})
    for brn, rows in seen_brn.items():
        if len(rows) > 1:
            issues.append({"type": "duplicate_business_registration_no", "sourceRows": rows, "businessRegistrationNo": brn})

    summary = {
        "count": len(companies),
        "uniqueLegacyIds": len(seen_ids),
        "uniqueBusinessRegistrationNos": len(seen_brn),
        "contractEntityTypes": Counter(normalize_text(c.get("contract_entity_type")) for c in companies if c.get("contract_entity_type")).most_common(),
    }
    return companies, issues, summary


def issue(issue_type: str, source_row: int, message: str, payload: dict[str, Any]) -> dict[str, Any]:
    return {
        "type": issue_type,
        "sourceRow": source_row,
        "message": message,
        "legacyKey": payload.get("legacy_key"),
        "legacyCompanyId": normalize_text(payload.get("legacy_company_id") or payload.get("headquarter_id")),
        "contractTitle": normalize_text(payload.get("contract_title")),
        "counterpartyName": normalize_text(payload.get("counterparty_name") or payload.get("company_name")),
    }


def sheet_stats(ws: Any) -> SheetStats:
    return SheetStats(rows=ws.max_row, cols=ws.max_column, nonempty=count_nonempty(ws))


def json_default(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat()
    if isinstance(value, Counter):
        return dict(value)
    return str(value)


def main() -> None:
    parser = argparse.ArgumentParser(description="Analyze legacy contract workbook without importing data.")
    parser.add_argument("--workbook", default=DEFAULT_WORKBOOK, help="Path to 계약현황.xlsm")
    parser.add_argument("--out", help="Optional JSON report output path")
    parser.add_argument("--pretty", action="store_true", help="Pretty-print JSON")
    args = parser.parse_args()

    workbook_path = Path(args.workbook)
    if not workbook_path.exists():
        raise SystemExit(f"Workbook not found: {workbook_path}")

    wb = load_workbook(workbook_path, read_only=True, keep_vba=True, data_only=True)
    required = [SHEET_CONTRACT_INPUT, SHEET_BILLING_LEDGER, SHEET_COMPANY]
    missing = [name for name in required if name not in wb.sheetnames]
    if missing:
        raise SystemExit("Required sheets missing: " + ", ".join(missing))

    contract_rows: list[dict[str, Any]] = []
    raw_contract_rows: list[tuple[Any, ...]] = []
    for source_row, row in iter_data_rows(wb[SHEET_CONTRACT_INPUT], 2):
        if normalize_text(get(row, CONTRACT_HEADERS["contract_title"])):
            item = row_to_contract(row, source_row)
            item["_row"] = row
            contract_rows.append(item)
            raw_contract_rows.append(row)

    contract_issues: list[dict[str, Any]] = []
    duplicate_keys: defaultdict[str, list[int]] = defaultdict(list)
    for item in contract_rows:
        duplicate_keys[item["legacy_key"]].append(item["source_row"])
        amount = number(item.get("order_amount"))
        if amount <= 0:
            contract_issues.append(issue("contract_amount", item["source_row"], "수주총액이 비어 있거나 0 이하입니다.", item))
        if not normalize_text(item.get("legacy_company_id")):
            contract_issues.append(issue("missing_company_id", item["source_row"], "업체ID가 비어 있습니다.", item))
        if not as_iso_date(item.get("contract_date")):
            contract_issues.append(issue("missing_contract_date", item["source_row"], "계약일이 비어 있거나 날짜로 인식되지 않습니다.", item))
    for key, rows in duplicate_keys.items():
        if len(rows) > 1:
            contract_issues.append(
                {
                    "type": "duplicate_contract_candidate",
                    "sourceRows": rows,
                    "legacyKey": key,
                    "message": "업체ID/계약명/계약일/금액 조합이 중복됩니다.",
                }
            )

    companies, company_issues, company_summary = analyze_companies(wb[SHEET_COMPANY])
    contract_summary = {
        "count": len(contract_rows),
        "uniqueCounterparties": len({normalize_company_name(c.get("counterparty_name")) for c in contract_rows if c.get("counterparty_name")}),
        "uniqueCompanyIds": len({normalize_text(c.get("legacy_company_id")) for c in contract_rows if c.get("legacy_company_id")}),
        "totalOrderAmount": round(sum(number(c.get("order_amount")) for c in contract_rows)),
        "serviceTypes": Counter(normalize_text(c.get("service_type")) for c in contract_rows if c.get("service_type")).most_common(),
        "contractYears": Counter((as_iso_date(c.get("contract_date")) or "")[:4] for c in contract_rows if as_iso_date(c.get("contract_date"))).most_common(),
    }

    report = {
        "workbook": {
            "path": str(workbook_path),
            "sizeBytes": workbook_path.stat().st_size,
            "modifiedAt": datetime.fromtimestamp(workbook_path.stat().st_mtime).isoformat(),
            "sheetCount": len(wb.sheetnames),
            "sheets": {
                name: sheet_stats(wb[name]).__dict__
                for name in [
                    SHEET_CONTRACT_INPUT,
                    SHEET_BILLING_LEDGER,
                    SHEET_COMPANY,
                    SHEET_CONTACT,
                    SHEET_FACILITY_DETAIL,
                ]
                if name in wb.sheetnames
            },
        },
        "contracts": contract_summary,
        "milestones": analyze_milestones(contract_rows),
        "billingLedger": analyze_billing_ledger(wb[SHEET_BILLING_LEDGER]),
        "companies": company_summary,
        "issues": {
            "count": len(contract_issues) + len(company_issues),
            "byType": Counter(item["type"] for item in [*contract_issues, *company_issues]).most_common(),
            "items": [*contract_issues, *company_issues][:500],
        },
        "migrationHints": {
            "primaryContractSheet": SHEET_CONTRACT_INPUT,
            "billingLedgerSheet": SHEET_BILLING_LEDGER,
            "companySheet": SHEET_COMPANY,
            "invoiceStoragePattern": r"매출계산서/{year}/{quarter}",
            "legacyKeyPolicy": "sha256(legacy_company_id|contract_title|contract_date|order_amount)[0:24]",
        },
    }

    text = json.dumps(report, ensure_ascii=False, indent=2 if args.pretty else None, default=json_default)
    if args.out:
        Path(args.out).write_text(text, encoding="utf-8")
    else:
        print(text)


if __name__ == "__main__":
    main()
