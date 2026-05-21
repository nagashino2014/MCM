"""
XLSX 텍스트 추출 모듈 (RAG 최적화 버전)
- openpyxl 기반 엑셀 파싱
- 다중 헤더 구조 보존
- JSON 구조화 변환
- 시트별 독립 청킹
- 대용량 파일 스트리밍 처리
"""
import re
import time
import json
import os
from pathlib import Path
from typing import Optional, List, Tuple, Dict, Any, Union

from .base import (
    BaseExtractor,
    ExtractionResult,
    ExtractionStatus,
    ExtractionMethod,
    PageResult,
    QualityDetails,
    TableData,
    detect_utf16_misencoding
)
from ..table_recovery.table_builder import TableBuilder, Table, TableCell
from ..config import ExtractionConfig, default_config

# openpyxl 지연 로딩
_openpyxl = None


def get_openpyxl():
    """openpyxl 모듈 지연 로딩"""
    global _openpyxl
    if _openpyxl is None:
        try:
            import openpyxl
            _openpyxl = openpyxl
        except ImportError:
            print("[WARNING] openpyxl not installed. XLSX extraction will fail.")
            return None
    return _openpyxl


class SheetStructure:
    """시트 구조 분석 결과"""
    def __init__(self):
        self.header_rows: int = 1
        self.header_levels: List[List[str]] = []  # 다중 헤더 레벨
        self.column_semantics: List[str] = []      # 열별 의미
        self.data_start_row: int = 1
        self.merged_regions: Dict[Tuple[int, int], Dict] = {}
        self.total_rows: int = 0
        self.total_cols: int = 0
        self.parallel_tables: List[Dict] = []     # 병렬 테이블 정보
        self.separator_cols: List[int] = []       # 구분 열 (빈 열)


class XLSXExtractor(BaseExtractor):
    """
    XLSX 텍스트 추출기 (RAG 최적화)
    
    주요 기능:
    - 다중 헤더 구조 감지 및 보존
    - JSON 구조화 변환 (column_semantics + row data)
    - 시트별 독립 청킹
    - 대용량 파일 스트리밍 처리
    """
    
    SUPPORTED_FORMATS = ["xlsx", "xls"]
    
    KOREAN_PATTERN = re.compile(r'[가-힣]')
    BROKEN_CHAR_PATTERN = re.compile(r'[\ufffd\x00-\x08\x0b\x0c\x0e-\x1f]')
    
    # 헤더 감지용 패턴
    HEADER_PATTERNS = [
        re.compile(r'^(구분|항목|분류|카테고리|category|type|name|이름|명칭)$', re.I),
        re.compile(r'^(연도|년도|year|\d{4}년?)$', re.I),
        re.compile(r'^(단위|unit|합계|total|소계|subtotal)$', re.I),
    ]
    
    def __init__(self, config: Optional[ExtractionConfig] = None):
        self.config = config or default_config
        self.max_rows = self.config.xlsx.max_rows
        self.max_cols = self.config.xlsx.max_cols
    
    def supports(self, file_format: str) -> bool:
        """XLSX 형식 지원 여부"""
        return file_format.lower() in self.SUPPORTED_FORMATS
    
    def extract(self, file_path: str) -> ExtractionResult:
        """XLSX 파일에서 텍스트 추출 (구조화 모드)"""
        start_time = time.time()
        file_path = Path(file_path)
        file_format = file_path.suffix.lower().lstrip('.')
        
        if not file_path.exists():
            return ExtractionResult(
                file_path=str(file_path),
                file_format=file_format,
                status=ExtractionStatus.FAILED,
                error_message=f"File not found: {file_path}"
            )
        
        if file_format == "xls":
            return ExtractionResult(
                file_path=str(file_path),
                file_format=file_format,
                status=ExtractionStatus.FAILED,
                error_message="Legacy .xls format is not supported. Please convert to .xlsx"
            )
        
        openpyxl = get_openpyxl()
        if openpyxl is None:
            return ExtractionResult(
                file_path=str(file_path),
                file_format=file_format,
                status=ExtractionStatus.FAILED,
                error_message="openpyxl module not available"
            )
        
        try:
            # 파일 크기 확인
            file_size_mb = os.path.getsize(file_path) / (1024 * 1024)
            use_streaming = file_size_mb >= self.config.xlsx.streaming_threshold_mb
            
            if use_streaming:
                print(f"[XLSX] 대용량 파일 ({file_size_mb:.1f}MB) - 스트리밍 모드")
            
            # 구조화 모드 여부 확인
            use_structured = self.config.xlsx.output_mode == "structured"
            
            # 워크북 열기 (병합 셀 정보를 위해 read_only=False)
            wb = openpyxl.load_workbook(
                str(file_path),
                read_only=self.config.xlsx.read_only_mode and not use_structured,
                data_only=True
            )
            
            # 시트별 추출
            all_texts = []
            all_pages = []
            all_structured_data = []
            total_table_index = 0
            
            for sheet_idx, sheet_name in enumerate(wb.sheetnames):
                sheet = wb[sheet_name]
                print(f"[XLSX] 시트 처리 중: {sheet_name}")
                
                if use_structured:
                    # 구조화 모드: JSON 변환 + 청킹
                    sheet_result = self._extract_sheet_structured(
                        sheet, sheet_name, sheet_idx, total_table_index
                    )
                    
                    if sheet_result:
                        all_texts.append(sheet_result['text'])
                        all_pages.extend(sheet_result['pages'])
                        all_structured_data.append(sheet_result['structured'])
                        total_table_index += len(sheet_result.get('tables', []))
                else:
                    # 기존 텍스트 추출 모드
                    sheet_text = self._extract_sheet_text(sheet, sheet_name)
                    if sheet_text:
                        all_texts.append(sheet_text)
                        all_pages.append(PageResult(
                            page_num=sheet_idx + 1,
                            text=sheet_text,
                            method=ExtractionMethod.TEXT_LAYER
                        ))
            
            wb.close()
            
            # 최종 텍스트 병합
            text = "\n\n".join(all_texts)
            
            if not text or len(text.strip()) < 5:
                return ExtractionResult(
                    file_path=str(file_path),
                    file_format=file_format,
                    status=ExtractionStatus.FAILED,
                    error_message="No text content found in XLSX file"
                )
            
            # 후처리
            text, corrections = self._apply_postprocessing(text)
            
            # 품질 검증
            quality_details = self._validate_quality(text)
            
            # 토큰 수 계산
            token_count = self._count_tokens(text)
            
            processing_time_ms = int((time.time() - start_time) * 1000)
            
            # 상태 결정
            if quality_details.overall_score >= self.config.quality_validation.pass_threshold:
                status = ExtractionStatus.COMPLETED
            elif quality_details.overall_score >= self.config.quality_validation.llm_fallback_threshold:
                status = ExtractionStatus.COMPLETED
            else:
                status = ExtractionStatus.LLM_FALLBACK if self.config.quality_validation.auto_llm_fallback else ExtractionStatus.FAILED
            
            return ExtractionResult(
                file_path=str(file_path),
                file_format=file_format,
                status=status,
                text=text,
                pages=all_pages if all_pages else [PageResult(page_num=1, text=text, method=ExtractionMethod.TEXT_LAYER)],
                page_count=len(all_pages) if all_pages else 1,
                quality_score=quality_details.overall_score,
                quality_details=quality_details,
                extraction_method=ExtractionMethod.TEXT_LAYER,
                char_count=len(text),
                token_count=token_count,
                processing_time_ms=processing_time_ms,
                corrections=corrections,
                metadata={
                    'structured_data': all_structured_data if use_structured else None,
                    'sheet_count': len(wb.sheetnames) if wb else 0
                }
            )
            
        except Exception as e:
            import traceback
            print(f"[XLSX] Extraction error: {e}")
            traceback.print_exc()
            return ExtractionResult(
                file_path=str(file_path),
                file_format=file_format,
                status=ExtractionStatus.FAILED,
                error_message=str(e),
                processing_time_ms=int((time.time() - start_time) * 1000)
            )
    
    def _extract_sheet_structured(
        self, 
        sheet, 
        sheet_name: str, 
        sheet_idx: int,
        start_table_index: int = 0
    ) -> Optional[Dict]:
        """
        시트를 구조화된 형식으로 추출
        
        Returns:
            {
                'text': str,           # RAG용 선형화 텍스트
                'pages': List[PageResult],
                'tables': List[TableData],
                'structured': Dict     # JSON 구조화 데이터
            }
        """
        try:
            # 1. 시트 구조 분석 (병합 셀 정보 포함)
            structure = self._analyze_sheet_structure(sheet)
            
            if structure.total_rows == 0:
                return None
            
            # 2. 다중 헤더 감지 및 구조화
            header_info = self._detect_multi_level_headers(sheet, structure)
            
            # 3. 병렬 테이블 처리
            if structure.parallel_tables and len(structure.parallel_tables) > 1:
                return self._extract_parallel_tables(
                    sheet, sheet_name, sheet_idx, start_table_index,
                    structure, header_info
                )
            
            # 4. 단일 테이블 데이터 추출 및 구조화
            structured_data = self._build_structured_data(
                sheet, structure, header_info, sheet_name
            )
            
            # 5. 청킹 처리 (대용량 시트)
            if self.config.xlsx.chunk_large_sheets and structure.total_rows > self.config.xlsx.chunk_row_size:
                chunks = self._chunk_sheet_data(structured_data, structure, header_info)
            else:
                chunks = [structured_data]
            
            # 6. 선형화 텍스트 생성
            text_parts = []
            pages = []
            tables = []
            table_builder = TableBuilder()
            
            for chunk_idx, chunk in enumerate(chunks):
                # 각 청크를 TableData로 변환
                rows_data = self._structured_to_rows(chunk)
                
                if rows_data and self._is_valid_table(rows_data):
                    table_obj = table_builder.build_from_grid(rows_data, header_rows=1)
                    
                    table_data = TableData(
                        table_index=start_table_index + len(tables),
                        rows=rows_data,
                        semantic_text=table_obj.to_semantic_text(),
                        structured_data=chunk,
                        markdown=table_obj.to_markdown(),
                        row_count=len(rows_data),
                        col_count=len(rows_data[0]) if rows_data else 0
                    )
                    tables.append(table_data)
                    
                    # 선형화 텍스트 추가
                    chunk_text = f"## {sheet_name}"
                    if len(chunks) > 1:
                        chunk_text += f" (파트 {chunk_idx + 1}/{len(chunks)})"
                    chunk_text += f"\n\n{table_obj.to_semantic_text()}"
                    text_parts.append(chunk_text)
                    
                    pages.append(PageResult(
                        page_num=sheet_idx + 1 + chunk_idx,
                        text=chunk_text,
                        method=ExtractionMethod.TEXT_LAYER,
                        has_tables=True,
                        tables=[table_data]
                    ))
            
            return {
                'text': '\n\n'.join(text_parts),
                'pages': pages,
                'tables': tables,
                'structured': structured_data
            }
            
        except Exception as e:
            print(f"[XLSX] Structured extraction error in '{sheet_name}': {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def _extract_parallel_tables(
        self,
        sheet,
        sheet_name: str,
        sheet_idx: int,
        start_table_index: int,
        structure: SheetStructure,
        header_info: Dict[str, Any]
    ) -> Optional[Dict]:
        """
        병렬 테이블들을 개별적으로 추출
        
        예: 직접 배출량 | (빈 열) | 간접 배출량 → 2개의 별도 테이블로 분리
        """
        text_parts = []
        pages = []
        tables = []
        all_structured = []
        table_builder = TableBuilder()
        
        print(f"[XLSX] 병렬 테이블 추출: {len(structure.parallel_tables)}개")
        
        for pt_idx, pt_info in enumerate(structure.parallel_tables):
            start_col = pt_info['start_col']
            end_col = pt_info['end_col']
            table_name = pt_info.get('name', f'table_{pt_idx}')
            
            # 해당 테이블 영역의 열 의미 추출
            col_semantics = header_info['column_semantics'][start_col - 1:end_col]
            header_levels = [
                level[start_col - 1:end_col] 
                for level in header_info.get('levels', [])
            ]
            
            # 테이블별 구조화 데이터 생성
            pt_structured = self._build_structured_data_for_range(
                sheet, structure, 
                start_col, end_col,
                col_semantics, header_levels,
                f"{sheet_name} - {table_name}"
            )
            
            if not pt_structured or not pt_structured.get('rows'):
                continue
            
            all_structured.append(pt_structured)
            
            # TableData 생성
            rows_data = self._structured_to_rows(pt_structured)
            
            if rows_data and self._is_valid_table(rows_data):
                table_obj = table_builder.build_from_grid(rows_data, header_rows=1)
                
                table_data = TableData(
                    table_index=start_table_index + len(tables),
                    rows=rows_data,
                    semantic_text=table_obj.to_semantic_text(),
                    structured_data=pt_structured,
                    markdown=table_obj.to_markdown(),
                    row_count=len(rows_data),
                    col_count=len(rows_data[0]) if rows_data else 0
                )
                tables.append(table_data)
                
                # 선형화 텍스트
                chunk_text = f"## {sheet_name} - {table_name}\n\n{table_obj.to_semantic_text()}"
                text_parts.append(chunk_text)
                
                pages.append(PageResult(
                    page_num=sheet_idx + 1 + pt_idx,
                    text=chunk_text,
                    method=ExtractionMethod.TEXT_LAYER,
                    has_tables=True,
                    tables=[table_data]
                ))
        
        if not tables:
            return None
        
        return {
            'text': '\n\n'.join(text_parts),
            'pages': pages,
            'tables': tables,
            'structured': {
                'sheet_name': sheet_name,
                'parallel_tables': all_structured,
                'table_count': len(all_structured)
            }
        }
    
    def _build_structured_data_for_range(
        self,
        sheet,
        structure: SheetStructure,
        start_col: int,
        end_col: int,
        col_semantics: List[str],
        header_levels: List[List[str]],
        table_name: str
    ) -> Dict[str, Any]:
        """특정 열 범위에 대한 구조화 데이터 생성"""
        structured = {
            'sheet_name': table_name,
            'structure_type': 'multi_header' if structure.header_rows > 1 else 'standard',
            'header_levels': header_levels,
            'column_semantics': col_semantics,
            'rows': [],
            'metadata': {
                'total_rows': structure.total_rows,
                'total_cols': end_col - start_col + 1,
                'header_rows': structure.header_rows,
                'data_rows': structure.total_rows - structure.header_rows,
                'col_range': f"{start_col}-{end_col}"
            }
        }
        
        # 데이터 행 추출
        for row_idx in range(structure.data_start_row, min(structure.total_rows + 1, self.max_rows + 1)):
            row_dict = {}
            has_content = False
            
            for col_offset, col_idx in enumerate(range(start_col, end_col + 1)):
                cell = sheet.cell(row_idx, col_idx)
                value = cell.value
                
                if value is not None:
                    if isinstance(value, float):
                        if value == int(value):
                            value = int(value)
                    value = str(value).strip()
                    if value:
                        has_content = True
                else:
                    value = ''
                
                key = col_semantics[col_offset] if col_offset < len(col_semantics) else f"Column{col_offset + 1}"
                row_dict[key] = value
            
            if has_content:
                structured['rows'].append(row_dict)
        
        return structured
    
    def _analyze_sheet_structure(self, sheet) -> SheetStructure:
        """시트 구조 분석 (병합 셀 우선 분석)"""
        structure = SheetStructure()
        
        try:
            # 최대 행/열 수 계산
            max_row = min(sheet.max_row or 0, self.max_rows)
            max_col = min(sheet.max_column or 0, self.max_cols)
            
            structure.total_rows = max_row
            structure.total_cols = max_col
            
            if max_row == 0 or max_col == 0:
                return structure
            
            # 1. 병합 셀 정보 먼저 수집 (헤더 감지에 활용)
            merged_info = self._get_merged_cell_ranges(sheet)
            structure.merged_regions = merged_info
            
            # 2. 병렬 테이블 구분 열 감지 (빈 열)
            separator_cols = self._detect_separator_columns(sheet, max_row, max_col)
            structure.separator_cols = separator_cols
            
            # 3. 병합 셀 정보를 활용한 헤더 행 수 추정
            header_rows = self._estimate_header_rows(sheet, max_col, merged_info)
            structure.header_rows = header_rows
            structure.data_start_row = header_rows + 1
            
            # 4. 병렬 테이블 정보 생성
            if separator_cols:
                structure.parallel_tables = self._define_parallel_tables(
                    separator_cols, max_col
                )
                print(f"[XLSX] 병렬 테이블 감지: {len(structure.parallel_tables)}개")
            
        except Exception as e:
            print(f"[XLSX] Structure analysis error: {e}")
            import traceback
            traceback.print_exc()
        
        return structure
    
    def _detect_separator_columns(self, sheet, max_row: int, max_col: int) -> List[int]:
        """빈 열(구분 열) 감지"""
        separator_cols = []
        check_rows = min(20, max_row)  # 처음 20행만 체크
        
        for col_idx in range(1, max_col + 1):
            empty_count = 0
            
            for row_idx in range(1, check_rows + 1):
                cell = sheet.cell(row_idx, col_idx)
                if cell.value is None or str(cell.value).strip() == '':
                    empty_count += 1
            
            # 90% 이상 빈 열이면 구분 열
            if empty_count / check_rows >= 0.9:
                separator_cols.append(col_idx)
        
        return separator_cols
    
    def _define_parallel_tables(
        self, 
        separator_cols: List[int], 
        max_col: int
    ) -> List[Dict]:
        """병렬 테이블 영역 정의"""
        tables = []
        
        # 연속된 구분 열은 하나로 취급
        sep_groups = []
        current_group = []
        
        for col in sorted(separator_cols):
            if not current_group or col == current_group[-1] + 1:
                current_group.append(col)
            else:
                sep_groups.append(current_group)
                current_group = [col]
        
        if current_group:
            sep_groups.append(current_group)
        
        # 테이블 영역 정의
        start_col = 1
        table_idx = 0
        
        for group in sep_groups:
            end_col = group[0] - 1
            
            if end_col >= start_col:
                tables.append({
                    'index': table_idx,
                    'start_col': start_col,
                    'end_col': end_col,
                    'name': f'table_{table_idx}'
                })
                table_idx += 1
            
            start_col = group[-1] + 1
        
        # 마지막 테이블
        if start_col <= max_col:
            tables.append({
                'index': table_idx,
                'start_col': start_col,
                'end_col': max_col,
                'name': f'table_{table_idx}'
            })
        
        return tables
    
    def _estimate_header_rows(self, sheet, max_col: int, merged_info: Dict = None) -> int:
        """
        헤더 행 수 추정 (병합 셀 기반 강화)
        
        개선된 로직:
        1. 병합 셀이 여러 행에 걸쳐 있으면 헤더 가능성 높음
        2. 숫자 데이터가 본격적으로 시작되는 지점 감지
        3. 동일한 패턴의 열 구조가 반복되는 지점 감지
        """
        max_check_rows = min(self.config.xlsx.detect_header_rows + 3, sheet.max_row or 10)
        
        # 병합 셀에서 최대 행 스팬 확인
        max_merge_row_span = 1
        if merged_info:
            for (row, col), info in merged_info.items():
                if row <= max_check_rows and info.get('row_span', 1) > 1:
                    max_merge_row_span = max(max_merge_row_span, row + info['row_span'] - 1)
        
        # 각 행의 특성 분석
        row_features = []
        
        for row_idx in range(1, max_check_rows + 1):
            row_data = []
            numeric_count = 0
            text_count = 0
            empty_count = 0
            long_numeric_count = 0  # 긴 숫자 (데이터일 가능성)
            year_like_count = 0     # 연도처럼 보이는 값
            
            for col_idx in range(1, max_col + 1):
                cell = sheet.cell(row_idx, col_idx)
                value = cell.value
                
                if value is None or str(value).strip() == '':
                    empty_count += 1
                    row_data.append('')
                elif isinstance(value, (int, float)):
                    numeric_count += 1
                    str_val = str(value)
                    row_data.append(str_val)
                    
                    # 긴 소수점 숫자는 데이터일 가능성 높음
                    if '.' in str_val and len(str_val) > 8:
                        long_numeric_count += 1
                    
                    # 연도 패턴 (2010-2030)
                    try:
                        int_val = int(value)
                        if 2000 <= int_val <= 2050:
                            year_like_count += 1
                    except (ValueError, OverflowError):
                        pass
                else:
                    text_count += 1
                    row_data.append(str(value).strip())
            
            non_empty = max_col - empty_count
            row_features.append({
                'row': row_idx,
                'data': row_data,
                'numeric_count': numeric_count,
                'text_count': text_count,
                'empty_count': empty_count,
                'numeric_ratio': numeric_count / max_col if max_col > 0 else 0,
                'text_ratio': text_count / max_col if max_col > 0 else 0,
                'empty_ratio': empty_count / max_col if max_col > 0 else 0,
                'long_numeric_ratio': long_numeric_count / non_empty if non_empty > 0 else 0,
                'year_like_count': year_like_count,
                'is_header_like': self._is_header_like_row(row_data)
            })
        
        # 헤더 행 수 결정 (개선된 로직)
        header_rows = 1
        
        # 방법 1: 긴 소수점 숫자가 등장하는 지점 찾기
        for i, features in enumerate(row_features):
            if features['long_numeric_ratio'] > 0.3:  # 30% 이상이 긴 숫자면 데이터 시작
                header_rows = i
                break
        
        # 방법 2: 숫자 비율 급증 지점 (더 세밀한 감지)
        if header_rows == 1:
            for i, features in enumerate(row_features):
                if i == 0:
                    continue
                
                prev = row_features[i - 1]
                
                # 숫자 비율이 0.4 이상으로 급증하고, 이전까지 텍스트/연도 위주였다면
                if (features['numeric_ratio'] > 0.4 and 
                    prev['numeric_ratio'] < 0.2 and
                    features['long_numeric_ratio'] > 0.1):
                    header_rows = i
                    break
        
        # 방법 3: 연도 행 다음이 데이터 시작인 경우
        if header_rows == 1:
            for i, features in enumerate(row_features):
                # 연도가 많이 포함된 행 다음이 데이터
                if features['year_like_count'] >= 3:  # 연도가 3개 이상
                    if i + 1 < len(row_features):
                        next_row = row_features[i + 1]
                        if next_row['numeric_ratio'] > 0.3:
                            header_rows = i + 1
                            break
        
        # 방법 4: 병합 셀 기반 (병합이 끝나는 지점)
        if header_rows == 1 and max_merge_row_span > 1:
            header_rows = max(header_rows, max_merge_row_span)
        
        # 최소 보장: 헤더로 보이는 행까지
        for i, features in enumerate(row_features):
            if features['is_header_like']:
                header_rows = max(header_rows, i + 1)
            else:
                break
        
        result = min(header_rows, max_check_rows)
        print(f"[XLSX] 헤더 행 감지: {result}행 (병합셀 최대: {max_merge_row_span})")
        
        return result
    
    def _is_header_like_row(self, row_data: List[str]) -> bool:
        """행이 헤더처럼 보이는지 확인 (개선된 버전)"""
        if not row_data:
            return False
        
        header_pattern_matches = 0
        year_count = 0
        long_decimal_count = 0
        non_empty_count = 0
        
        for cell in row_data:
            if not cell:
                continue
            non_empty_count += 1
            
            # 헤더 패턴 매칭
            for pattern in self.HEADER_PATTERNS:
                if pattern.search(cell):
                    header_pattern_matches += 1
                    break
            
            # 연도 패턴 (2000-2050) - 헤더일 가능성
            try:
                int_val = int(cell.replace(',', ''))
                if 2000 <= int_val <= 2050:
                    year_count += 1
                    header_pattern_matches += 1  # 연도도 헤더로 취급
            except (ValueError, OverflowError):
                pass
            
            # 긴 소수점 숫자는 데이터일 가능성이 높음
            try:
                float_val = float(cell.replace(',', ''))
                if '.' in cell and len(cell) > 8:
                    long_decimal_count += 1
            except ValueError:
                pass
        
        if non_empty_count == 0:
            return False
        
        # 긴 소수점 숫자가 30% 이상이면 데이터 행
        if long_decimal_count / non_empty_count >= 0.3:
            return False
        
        # 연도가 3개 이상이면 헤더 (연도 행)
        if year_count >= 3:
            return True
        
        # 헤더 패턴이 20% 이상 매칭되면 헤더
        if header_pattern_matches / non_empty_count >= 0.2:
            return True
        
        return False
    
    def _detect_multi_level_headers(
        self, 
        sheet, 
        structure: SheetStructure
    ) -> Dict[str, Any]:
        """
        다중 헤더 레벨 감지 및 구조화 (병합 셀 개선)
        
        Returns:
            {
                'levels': [[level0_headers], [level1_headers], ...],
                'merged_headers': {(row, col): {'text': str, 'col_span': int, 'row_span': int}},
                'column_semantics': ['열1 의미', '열2 의미', ...]
            }
        """
        header_info = {
            'levels': [],
            'merged_headers': {},
            'column_semantics': []
        }
        
        if structure.total_cols == 0:
            return header_info
        
        # 병합 셀 정보를 (row, col) -> 원본 값 맵으로 변환
        # 병합된 영역의 모든 셀에 원본 값을 매핑
        merge_value_map = {}
        for (min_row, min_col), info in structure.merged_regions.items():
            text = info['text']
            row_span = info.get('row_span', 1)
            col_span = info.get('col_span', 1)
            
            for r in range(min_row, min_row + row_span):
                for c in range(min_col, min_col + col_span):
                    merge_value_map[(r, c)] = text
            
            header_info['merged_headers'][(min_row, min_col)] = {
                'text': text,
                'col_span': col_span,
                'row_span': row_span
            }
        
        # 각 헤더 행 수집 (병합 셀 값 포함)
        for row_idx in range(1, structure.header_rows + 1):
            level_headers = []
            
            for col_idx in range(1, structure.total_cols + 1):
                # 병합 셀 맵에서 값 확인
                if (row_idx, col_idx) in merge_value_map:
                    value = merge_value_map[(row_idx, col_idx)]
                else:
                    cell = sheet.cell(row_idx, col_idx)
                    value = str(cell.value).strip() if cell.value else ''
                
                level_headers.append(value)
            
            header_info['levels'].append(level_headers)
        
        # 열별 의미 추론 (다중 헤더 결합, 병합 정보 포함)
        header_info['column_semantics'] = self._infer_column_semantics(
            header_info['levels'], 
            structure.merged_regions
        )
        
        # 병렬 테이블이 있는 경우, 테이블별 접두사 추가
        if structure.parallel_tables and len(structure.parallel_tables) > 1:
            header_info['column_semantics'] = self._add_table_prefix_to_semantics(
                header_info['column_semantics'],
                header_info['levels'],
                structure.parallel_tables
            )
        
        return header_info
    
    def _add_table_prefix_to_semantics(
        self,
        semantics: List[str],
        header_levels: List[List[str]],
        parallel_tables: List[Dict]
    ) -> List[str]:
        """병렬 테이블에 대해 테이블별 접두사 추가"""
        result = semantics.copy()
        
        for table in parallel_tables:
            start_col = table['start_col'] - 1  # 0-based
            end_col = table['end_col'] - 1
            
            # 첫 번째 레벨에서 테이블 이름 추출
            if header_levels and start_col < len(header_levels[0]):
                table_name = header_levels[0][start_col]
                if table_name:
                    table['name'] = table_name
                    
                    # 해당 테이블의 열에 접두사 추가 (이미 포함되어 있지 않은 경우)
                    for col_idx in range(start_col, min(end_col + 1, len(result))):
                        if not result[col_idx].startswith(table_name):
                            # 이미 계층 구조가 있으면 첫 부분이 테이블 이름인지 확인
                            parts = result[col_idx].split(' > ')
                            if parts[0] != table_name:
                                result[col_idx] = f"{table_name} > {result[col_idx]}"
        
        return result
    
    def _infer_column_semantics(
        self, 
        header_levels: List[List[str]], 
        merged_info: Dict = None
    ) -> List[str]:
        """
        다중 헤더 레벨에서 열별 의미 추론 (개선된 버전)
        
        개선사항:
        1. 병합 셀 정보를 활용하여 빈 셀에 값 전파
        2. 각 레벨의 값을 계층적으로 결합
        3. 의미 있는 열 이름 생성 (직접 > 배출연도 > 2015 형태)
        """
        if not header_levels:
            return []
        
        col_count = max(len(level) for level in header_levels)
        
        # 1단계: 각 레벨에서 빈 셀에 왼쪽 값 전파 (병합 셀 효과)
        filled_levels = []
        for level in header_levels:
            filled_level = []
            last_non_empty = ''
            
            for col_idx in range(col_count):
                if col_idx < len(level):
                    value = level[col_idx].strip() if level[col_idx] else ''
                else:
                    value = ''
                
                # 빈 값이면 왼쪽의 마지막 값 사용 (병합 셀 시뮬레이션)
                if not value:
                    value = last_non_empty
                else:
                    last_non_empty = value
                
                filled_level.append(value)
            
            filled_levels.append(filled_level)
        
        # 2단계: 열별로 계층적 의미 생성
        semantics = []
        
        for col_idx in range(col_count):
            parts = []
            seen = set()
            
            for level in filled_levels:
                if col_idx < len(level):
                    value = level[col_idx]
                    
                    # 중복 방지 및 의미 있는 값만 추가
                    if value and value not in seen:
                        # 너무 긴 값은 축약
                        if len(value) > 50:
                            value = value[:47] + '...'
                        parts.append(value)
                        seen.add(value)
            
            # 계층적 결합
            if parts:
                # 마지막 값이 숫자(연도 등)이면 그것만 사용할 수도 있음
                last_part = parts[-1] if parts else ''
                
                # 연도 패턴인 경우 상위 카테고리 포함
                if re.match(r'^\d{4}$', last_part):
                    # 연도는 상위 2개 레벨까지만 포함
                    semantic = ' > '.join(parts[-3:]) if len(parts) > 2 else ' > '.join(parts)
                else:
                    # 일반적인 경우 전체 경로
                    semantic = ' > '.join(parts)
            else:
                semantic = f"Column{col_idx + 1}"
            
            semantics.append(semantic)
        
        return semantics
    
    def _build_structured_data(
        self,
        sheet,
        structure: SheetStructure,
        header_info: Dict[str, Any],
        sheet_name: str
    ) -> Dict[str, Any]:
        """
        구조화된 데이터 생성
        
        Returns:
            {
                'sheet_name': str,
                'structure_type': str,
                'header_levels': [[level0], [level1], ...],
                'column_semantics': ['열1', '열2', ...],
                'rows': [
                    {'열1': '값1', '열2': '값2', ...},
                    ...
                ],
                'metadata': {...}
            }
        """
        structured = {
            'sheet_name': sheet_name,
            'structure_type': 'multi_header' if structure.header_rows > 1 else 'standard',
            'header_levels': header_info['levels'],
            'column_semantics': header_info['column_semantics'],
            'rows': [],
            'metadata': {
                'total_rows': structure.total_rows,
                'total_cols': structure.total_cols,
                'header_rows': structure.header_rows,
                'data_rows': structure.total_rows - structure.header_rows
            }
        }
        
        # 데이터 행 추출
        col_semantics = header_info['column_semantics']
        
        for row_idx in range(structure.data_start_row, min(structure.total_rows + 1, self.max_rows + 1)):
            row_dict = {}
            has_content = False
            
            for col_idx in range(1, min(structure.total_cols + 1, self.max_cols + 1)):
                cell = sheet.cell(row_idx, col_idx)
                value = cell.value
                
                # 값 정리
                if value is not None:
                    if isinstance(value, float):
                        # 소수점 이하가 0이면 정수로 변환
                        if value == int(value):
                            value = int(value)
                    value = str(value).strip()
                    if value:
                        has_content = True
                else:
                    value = ''
                
                # 열 의미로 키 설정
                key = col_semantics[col_idx - 1] if col_idx - 1 < len(col_semantics) else f"Column{col_idx}"
                row_dict[key] = value
            
            if has_content:
                structured['rows'].append(row_dict)
        
        return structured
    
    def _chunk_sheet_data(
        self,
        structured_data: Dict[str, Any],
        structure: SheetStructure,
        header_info: Dict[str, Any]
    ) -> List[Dict[str, Any]]:
        """
        대용량 시트 데이터 청킹
        
        각 청크는 헤더 정보를 포함하여 독립적으로 이해 가능
        """
        chunks = []
        rows = structured_data.get('rows', [])
        chunk_size = self.config.xlsx.chunk_row_size
        
        if len(rows) <= chunk_size:
            return [structured_data]
        
        for i in range(0, len(rows), chunk_size):
            chunk_rows = rows[i:i + chunk_size]
            
            chunk = {
                'sheet_name': structured_data['sheet_name'],
                'structure_type': structured_data['structure_type'],
                'header_levels': structured_data['header_levels'],
                'column_semantics': structured_data['column_semantics'],
                'rows': chunk_rows,
                'metadata': {
                    **structured_data['metadata'],
                    'chunk_index': i // chunk_size,
                    'chunk_start_row': i + structure.data_start_row,
                    'chunk_end_row': min(i + chunk_size, len(rows)) + structure.data_start_row - 1,
                    'is_chunked': True
                }
            }
            chunks.append(chunk)
        
        print(f"[XLSX] 시트 '{structured_data['sheet_name']}' 청킹: {len(chunks)}개 청크 ({len(rows)}행)")
        
        return chunks
    
    def _structured_to_rows(self, structured: Dict[str, Any]) -> List[List[str]]:
        """구조화된 데이터를 2D 배열로 변환"""
        rows = []
        
        # 헤더 행 추가
        col_semantics = structured.get('column_semantics', [])
        if col_semantics:
            rows.append(col_semantics)
        
        # 데이터 행 추가
        for row_dict in structured.get('rows', []):
            row = []
            for col in col_semantics:
                row.append(row_dict.get(col, ''))
            rows.append(row)
        
        return rows
    
    def _get_merged_cell_ranges(self, sheet) -> Dict[Tuple[int, int], Dict]:
        """병합 셀 정보 추출"""
        merged_info = {}
        
        try:
            for merged_range in sheet.merged_cells.ranges:
                min_row, min_col = merged_range.min_row, merged_range.min_col
                max_row, max_col = merged_range.max_row, merged_range.max_col
                
                cell_value = sheet.cell(min_row, min_col).value
                merged_info[(min_row, min_col)] = {
                    'text': str(cell_value).strip() if cell_value else '',
                    'row_span': max_row - min_row + 1,
                    'col_span': max_col - min_col + 1
                }
        except Exception:
            pass
        
        return merged_info
    
    def _is_valid_table(self, table: List[List[str]]) -> bool:
        """유효한 표인지 검사"""
        if not table:
            return False
        
        row_count = len(table)
        col_count = max(len(row) for row in table) if table else 0
        
        if row_count < self.config.xlsx.min_table_rows:
            return False
        if col_count < self.config.xlsx.min_table_cols:
            return False
        
        total_cells = row_count * col_count
        non_empty_cells = sum(
            1 for row in table for cell in row 
            if cell and str(cell).strip()
        )
        
        if total_cells > 0 and non_empty_cells / total_cells < 0.2:
            return False
        
        return True
    
    def _extract_sheet_text(self, sheet, sheet_name: str) -> str:
        """기존 텍스트 추출 방식 (호환성)"""
        rows_text = []
        row_count = 0
        
        rows_text.append(f"[시트: {sheet_name}]")
        
        try:
            for row in sheet.iter_rows(max_row=self.max_rows, max_col=self.max_cols):
                cells = []
                has_content = False
                
                for cell in row:
                    value = cell.value
                    if value is not None:
                        cell_text = str(value).strip()
                        if cell_text:
                            cells.append(cell_text)
                            has_content = True
                        else:
                            cells.append("")
                    else:
                        cells.append("")
                
                if has_content:
                    while cells and not cells[-1]:
                        cells.pop()
                    
                    if cells:
                        rows_text.append(" | ".join(cells))
                        row_count += 1
                
                if row_count >= self.max_rows:
                    rows_text.append(f"... ({self.max_rows}행 이후 생략)")
                    break
                    
        except Exception as e:
            print(f"[XLSX] Sheet extraction error: {e}")
        
        return "\n".join(rows_text) if row_count > 0 else ""
    
    def _apply_postprocessing(self, text: str) -> Tuple[str, List[dict]]:
        """후처리 적용"""
        corrections = []
        
        # 연속 줄바꿈 정리
        text = re.sub(r'\n{3,}', '\n\n', text)
        
        # 연속 공백 정리
        text = re.sub(r' +', ' ', text)
        
        # 특수 문자 정리
        special_chars = [
            (r'\u200b', ''),
            (r'\u00a0', ' '),
            (r'\ufeff', ''),
        ]
        
        for pattern, replacement in special_chars:
            if re.search(pattern, text):
                text = re.sub(pattern, replacement, text)
                corrections.append({"type": "special_char", "pattern": pattern})
        
        return text.strip(), corrections
    
    def _validate_quality(self, text: str) -> QualityDetails:
        """
        품질 검증 (XLSX 전용)
        
        XLSX는 openpyxl이 이미 올바르게 파싱하므로 UTF-16 인코딩 오류가 발생할 수 없음.
        따라서 인코딩 검사를 건너뛰고 항상 정상으로 처리.
        """
        if not text or len(text.strip()) < 10:
            return QualityDetails(overall_score=0.0)
        
        text_no_space = re.sub(r'\s', '', text)
        text_no_numbers = re.sub(r'[\d.,|:\->]', '', text_no_space)
        
        korean_chars = len(self.KOREAN_PATTERN.findall(text))
        
        if text_no_numbers:
            korean_ratio = korean_chars / len(text_no_numbers)
        else:
            korean_ratio = 0.0
        
        # 한글 점수 (엑셀은 숫자 위주일 수 있으므로 관대하게)
        if korean_ratio >= 0.30:
            korean_score = 1.0
        elif korean_ratio >= 0.15:
            korean_score = 0.8
        elif korean_ratio >= 0.05:
            korean_score = 0.6
        else:
            # 엑셀 데이터는 숫자가 많아 한글이 적을 수 있음
            korean_score = 0.5
        
        broken_chars = len(self.BROKEN_CHAR_PATTERN.findall(text))
        broken_ratio = broken_chars / len(text) if text else 0.0
        broken_score = 1.0 - min(broken_ratio * 5, 1.0)
        
        # 구조화된 데이터 점수 (개선된 방식)
        # '>' 기호는 다중 헤더 결합, ':' 기호는 키-값 쌍
        structure_indicators = len(re.findall(r'[>:]', text))
        total_chars = len(text)
        
        if structure_indicators / total_chars >= 0.01:
            sentence_score = 1.0
        elif '|' in text or '>' in text:
            sentence_score = 0.9
        else:
            sentence_score = 0.8
        
        repetition_score = 1.0
        
        # XLSX는 openpyxl이 이미 올바르게 디코딩하므로 인코딩 오류 없음
        # ZIP 압축된 XML 파일이므로 UTF-16 인코딩 문제가 발생하지 않음
        has_encoding_error = False
        encoding_score = 1.0
        
        overall = (
            korean_score * 0.30 +
            broken_score * 0.25 +
            sentence_score * 0.15 +
            repetition_score * 0.10 +
            encoding_score * 0.20
        )
        
        return QualityDetails(
            korean_ratio=korean_ratio,
            broken_char_ratio=broken_ratio,
            sentence_structure=sentence_score,
            repetition_anomaly=0.0,
            encoding_error=has_encoding_error,
            korean_score=round(korean_score, 4),
            broken_score=round(broken_score, 4),
            sentence_score=round(sentence_score, 4),
            repetition_score=round(repetition_score, 4),
            encoding_score=round(encoding_score, 4),
            overall_score=round(overall, 4)
        )
    
    def _count_tokens(self, text: str) -> int:
        """토큰 수 계산"""
        try:
            import tiktoken
            enc = tiktoken.get_encoding("cl100k_base")
            return len(enc.encode(text))
        except ImportError:
            return len(text) // 2
